
import streamlit as st
import pandas as pd
import numpy as np
import json, math
from io import BytesIO
from openpyxl import Workbook
from openpyxl.chart import LineChart, Reference
from openpyxl.styles import Font

st.set_page_config(page_title="LifePlan v4", layout="wide")

# -----------------------------
# Financial / pension helpers
# -----------------------------
def progressive_tax(taxable, brackets):
    tax = 0.0
    lower = 0.0
    for upper, rate in brackets:
        slice_amt = max(0.0, min(taxable, upper) - lower)
        tax += slice_amt * rate
        lower = upper
        if taxable <= upper:
            break
    if taxable > lower:
        tax += (taxable-lower) * brackets[-1][1]
    return max(0.0, tax)

def net_salary_from_gross(gross, worker_contrib_rate, regional_local_rate,
                          irpef_1, irpef_2, irpef_3):
    # Cash-pay estimate, distinct from INPS pension "aliquota di computo".
    worker_contrib = gross * worker_contrib_rate
    taxable = max(0.0, gross - worker_contrib)
    brackets = [(28000, irpef_1), (50000, irpef_2), (10**12, irpef_3)]
    irpef = progressive_tax(taxable, brackets)
    local = taxable * regional_local_rate
    return max(0.0, gross - worker_contrib - irpef - local)

def annuity_payment(principal, annual_rate, years):
    if principal <= 0 or years <= 0:
        return 0.0
    r = annual_rate / 12
    n = years * 12
    if abs(r) < 1e-12:
        return principal / n
    return principal * r / (1 - (1+r)**(-n))

def mortgage_balance(P, annual_rate, years, elapsed_years):
    if elapsed_years >= years:
        return 0.0
    r = annual_rate/12
    n = years*12
    m = max(0, elapsed_years*12)
    if abs(r) < 1e-12:
        return max(0.0, P*(1-m/n))
    pmt = P*r/(1-(1+r)**(-n))
    return max(0.0, P*(1+r)**m - pmt*((1+r)**m-1)/r)

# Current coefficients are parameterized via a simple table so they can be replaced later.
# Approximate 2025-26 coefficients for planning, not an INPS certificate.
DEFAULT_COEFF = {
    57:0.04204,58:0.04308,59:0.04419,60:0.04536,61:0.04661,62:0.04795,
    63:0.04936,64:0.05088,65:0.05250,66:0.05423,67:0.05608,68:0.05808,
    69:0.06024,70:0.06258,71:0.06510
}
def transformation_coeff(age):
    if age in DEFAULT_COEFF:
        return DEFAULT_COEFF[age]
    if age < 57:
        return DEFAULT_COEFF[57]
    return DEFAULT_COEFF[71]

def pension_net_from_gross(gross, regional_local_rate, irpef_1, irpef_2, irpef_3):
    taxable = gross
    brackets = [(28000, irpef_1), (50000, irpef_2), (10**12, irpef_3)]
    return max(0.0, gross - progressive_tax(taxable, brackets) - taxable*regional_local_rate)

def apply_black_swans(invested, age, swans):
    notes = []
    for s in swans:
        if not s["active"]:
            continue
        start = s["age"]
        dur = s["duration"]
        if age == start:
            invested *= max(0.0, 1 - s["crash"])
            notes.append(f'Cigno nero: -{s["crash"]*100:.0f}%')
        elif start < age < start + dur:
            invested *= (1 + s["crisis_return"])
            notes.append(f'Cigno nero: rendimento crisi {s["crisis_return"]*100:.1f}%')
    return invested, "; ".join(notes)

def risk_label(df, target, retirement_age):
    min_money = df["Patrimonio monetario"].min()
    money90 = df.iloc[-1]["Patrimonio monetario"]
    neg_years = int((df["Patrimonio monetario"] <= 0).sum())
    dec_rows = df[df["Flusso netto dopo abitazione"] < 0]
    dec_age = None if dec_rows.empty else int(dec_rows.iloc[0]["Età"])
    score = 0
    if neg_years > 0: score += 5
    if money90 < target: score += 3
    elif money90 < target*2: score += 2
    elif money90 < target*4: score += 1
    if min_money < target: score += 2
    if dec_age is not None and dec_age < retirement_age: score += 2
    if score >= 7: return "Alto"
    if score >= 4: return "Medio-alto"
    if score >= 2: return "Medio"
    return "Basso"


def gross_salary_for_age(p, age):
    """RAL base (escluso bonus) secondo la modalità carriera selezionata."""
    age0 = int(p["age_start"])
    if age < age0:
        return 0.0
    mode = p.get("salary_mode", "Crescita costante")
    g = float(p["gross_salary"])
    growth = float(p["salary_growth"])

    if mode == "Crescita costante":
        return g * ((1 + growth) ** (age - age0))

    if mode == "Scatti di carriera":
        current = g
        steps = sorted([s for s in p.get("salary_steps", []) if s.get("active", False)],
                       key=lambda x: x["age"])
        for a in range(age0, age + 1):
            if a > age0:
                current *= (1 + growth)
            for s in steps:
                if int(s["age"]) == a:
                    current = float(s["gross"])
        return current

    pts = sorted(
        [(int(x["age"]), float(x["gross"])) for x in p.get("salary_custom_points", [])
         if float(x.get("gross", 0)) > 0],
        key=lambda x: x[0]
    )
    if not pts:
        return g * ((1 + growth) ** (age - age0))
    if age <= pts[0][0]:
        return pts[0][1] / ((1 + growth) ** max(0, pts[0][0] - age))
    for (a1, g1), (a2, g2) in zip(pts[:-1], pts[1:]):
        if a1 <= age <= a2:
            if a2 == a1:
                return g2
            w = (age - a1) / (a2 - a1)
            return g1 + (g2 - g1) * w
    last_age, last_gross = pts[-1]
    return last_gross * ((1 + growth) ** max(0, age - last_age))

def annual_bonus_for_age(p, age):
    if age >= int(p["retire_age"]):
        return 0.0
    bonus = float(p.get("annual_bonus", 0.0))
    if p.get("bonus_growth_with_salary", True):
        bonus *= (1 + float(p["salary_growth"])) ** max(0, age - int(p["age_start"]))
    return bonus

# -----------------------------
# Core simulation
# -----------------------------
def simulate(p, housing, down_pct=0.0, lifecycle="keep"):
    age0 = int(p["age_start"]); age_end = int(p["age_end"])
    retire_age = int(p["retire_age"])
    inps_age = int(p["inps_age"])
    derisk_age = int(p["derisk_age"])

    # initial monetary wealth split
    invest_share = p["invest_share"]
    money0 = p["initial_wealth"]
    invested = money0 * invest_share
    liquid = money0 * (1-invest_share)

    house = 0.0
    debt = 0.0
    mortgage_principal = 0.0
    mortgage_pmt_m = 0.0
    special_nuda = 0.0
    nuda_done = False
    downsized = False

    # Initial INPS + complementary fund
    inps_montante = p["inps_initial_montante"]
    fund = p["fund_initial"]
    years_contrib_initial = p["contrib_years_initial"]
    years_fund_initial = p["fund_years_initial"]

    if housing == "owned":
        house = p["owned_home_value"]
        debt = 0.0
    elif housing == "buy":
        down = p["house_price"] * down_pct
        purchase_cash = down + p["buying_costs"]
        # fund from liquid first, then investments
        use_liq = min(liquid, purchase_cash)
        liquid -= use_liq
        remain = purchase_cash - use_liq
        invested = max(0.0, invested-remain)
        house = p["house_price"]
        mortgage_principal = p["house_price"] - down
        debt = mortgage_principal
        mortgage_pmt_m = annuity_payment(mortgage_principal, p["mortgage_rate"], p["mortgage_years"])

    rows = []
    # RITA standard eligibility: within 5 years of INPS age, >=20 contrib years, >=5 fund years.
    # If retirement before that, bridge uses monetary wealth until eligible.
    rita_start_age = max(retire_age, inps_age-5)
    rita_active = False
    rita_remaining = 0.0

    # Pension determined at INPS start from montante reached then
    inps_pension_gross = 0.0
    inps_pension_net = 0.0
    inps_locked = False

    for age in range(age0, age_end+1):
        y = age-age0
        base_gross = gross_salary_for_age(p, age) if age < retire_age else 0.0
        bonus_gross = annual_bonus_for_age(p, age) if age < retire_age else 0.0
        gross = base_gross + bonus_gross
        living = p["living_monthly"]*12*((1+p["inflation"])**y)
        owner_cost = p["owner_cost_monthly"]*12*((1+p["inflation"])**y)
        rent = p["rent_monthly"]*12*((1+p["rent_growth"])**y)

        # House value evolves nominally
        if housing in ("buy","owned") and y > 0 and house > 0:
            house *= (1+p["house_app"])

        # salary / contributions
        salary_net = 0.0
        tfr_contribution = 0.0
        if age < retire_age:
            salary_net = net_salary_from_gross(
                gross, p["worker_contrib_rate"], p["regional_local_rate"],
                p["irpef1"], p["irpef2"], p["irpef3"]
            )
            # INPS montante: contribution base × computo rate, then annual capitalization.
            inps_montante *= (1+p["inps_revaluation"])
            inps_montante += gross*p["inps_computo_rate"]
            years_contrib = years_contrib_initial + (age-age0+1)

            # TFR to complementary fund, approximated as gross/13.5
            tfr_contribution = gross / 13.5
            fund_return = p["return_pre"] if age < derisk_age else p["return_post"]
            fund *= (1+fund_return)
            fund += tfr_contribution
            years_fund = years_fund_initial + (age-age0+1)
        else:
            years_contrib = years_contrib_initial + max(0, retire_age-age0)
            years_fund = years_fund_initial + max(0, retire_age-age0)
            # position continues to earn until RITA starts
            if not rita_active and fund > 0:
                fund_return = p["return_pre"] if age < derisk_age else p["return_post"]
                fund *= (1+fund_return)

        # INPS starts at chosen minimum pension age
        if age >= inps_age and not inps_locked:
            coeff = transformation_coeff(inps_age)
            inps_pension_gross = inps_montante*coeff
            inps_pension_net = pension_net_from_gross(
                inps_pension_gross, p["regional_local_rate"], p["irpef1"], p["irpef2"], p["irpef3"]
            )
            inps_locked = True

        # RITA: standard route; paid pro-rata until INPS age and exhausts fund.
        rita_net = 0.0
        standard_rita_ok = (
            age >= rita_start_age and age < inps_age and
            years_contrib >= 20 and years_fund >= 5 and age >= retire_age
        )
        if standard_rita_ok and fund > 0:
            if not rita_active:
                rita_active = True
                rita_remaining = fund
            years_left = max(1, inps_age-age)
            rita_gross = fund / years_left
            # Simplified favorable taxation as adjustable rate
            rita_net = rita_gross*(1-p["rita_tax_rate"])
            fund = max(0.0, fund-rita_gross)

        # income
        inps_net_this_year = inps_pension_net if age >= inps_age else 0.0
        income = salary_net + rita_net + inps_net_this_year

        # housing
        if housing == "rent":
            housing_cost = rent
            mortgage_annual = 0.0
        elif housing == "owned":
            mortgage_annual = 0.0
            debt = 0.0
            housing_cost = (
                p["owned_home_ordinary_monthly"]*12*((1+p["inflation"])**y)
                + p["owned_home_extra_annual"]*((1+p["inflation"])**y)
            )
        else:
            elapsed = age-age0
            debt = mortgage_balance(mortgage_principal, p["mortgage_rate"], p["mortgage_years"], elapsed)
            mortgage_annual = mortgage_pmt_m*12 if elapsed < p["mortgage_years"] else 0.0
            housing_cost = mortgage_annual + owner_cost

            # nuda property
            if lifecycle == "nuda" and (not nuda_done) and age >= p["nuda_age"] and house > 0:
                proceeds = house*p["nuda_pct"]
                special_nuda += proceeds
                house = 0.0
                nuda_done = True

            # downsizing
            if lifecycle == "downsize" and (not downsized) and age >= p["downsize_age"] and house > 0:
                smaller = p["smaller_house_today"]*((1+p["house_app"])**y)
                released = max(0.0, house-smaller-p["downsize_costs"])
                # released money split by current investment share
                invested += released*p["invest_share"]
                liquid += released*(1-p["invest_share"])
                house = smaller
                debt = 0.0
                mortgage_annual = 0.0
                owner_cost = p["smaller_owner_cost_monthly"]*12*((1+p["inflation"])**y)
                housing_cost = owner_cost
                downsized = True

        cashflow = income - living - housing_cost

        # Returns, black swans only on risky/invested portfolio
        base_return = p["return_pre"] if age < derisk_age else p["return_post"]
        invested *= (1+base_return)
        invested, swan_note = apply_black_swans(invested, age, p["swans"])

        # nuda proceeds have own low-risk return
        special_nuda *= (1+p["nuda_return"])

        # Savings / decumulation
        if cashflow >= 0:
            invested += cashflow*p["saving_invest_share"]
            liquid += cashflow*(1-p["saving_invest_share"])
        else:
            need = -cashflow
            # use nuda reserve, then liquid, then invested
            u = min(special_nuda, need); special_nuda -= u; need -= u
            u = min(liquid, need); liquid -= u; need -= u
            if need > 0:
                invested -= need

        monetary = invested + liquid + special_nuda + fund
        total = monetary + house - debt

        rows.append({
            "Età":age,
            "Fase":"Lavoro" if age < retire_age else ("Bridge/RITA" if age < inps_age else "Pensione INPS"),
            "RAL base":base_gross,
            "Bonus lordo":bonus_gross,
            "RAL lorda complessiva":gross,
            "Reddito netto lavoro":salary_net,
            "RITA netta":rita_net,
            "Pensione INPS netta":inps_net_this_year,
            "Reddito totale":income,
            "Spese personali":living,
            "Costo abitazione":housing_cost,
            "Flusso netto dopo abitazione":cashflow,
            "Patrimonio investito":invested,
            "Liquidità":liquid,
            "Riserva nuda proprietà":special_nuda,
            "Fondo pensione":fund,
            "Montante INPS":inps_montante,
            "Valore casa":house,
            "Debito residuo":debt,
            "Patrimonio monetario":monetary,
            "Patrimonio totale":total,
            "Evento mercato":swan_note
        })
    return pd.DataFrame(rows)

def summarize(df, p, label):
    dec = df[df["Flusso netto dopo abitazione"] < 0]
    return {
        "Scenario":label,
        "Patrimonio monetario a 90":float(df.iloc[-1]["Patrimonio monetario"]),
        "Valore casa a 90":float(df.iloc[-1]["Valore casa"]),
        "Patrimonio totale a 90":float(df.iloc[-1]["Patrimonio totale"]),
        "Patrimonio monetario minimo":float(df["Patrimonio monetario"].min()),
        "Inizio decumulo":"-" if dec.empty else int(dec.iloc[0]["Età"]),
        "Pensione INPS netta annua":float(df.iloc[-1]["Pensione INPS netta"]),
        "Rischio":risk_label(df,p["target_final"],p["retire_age"]),
        "Sostenibile":"Sì" if (df["Patrimonio monetario"].min()>=0 and df.iloc[-1]["Patrimonio monetario"]>=p["target_final"]) else "No"
    }

def excel_report(p, results, summary):
    bio=BytesIO(); wb=Workbook(); ws=wb.active; ws.title="Riepilogo"
    ws["A1"]="LifePlan v4 - Input e output"; ws["A1"].font=Font(bold=True,size=14)
    r=3; ws.cell(r,1,"INPUT"); ws.cell(r,1).font=Font(bold=True); r+=1
    for k,v in p.items():
        if k=="swans":
            for i,s in enumerate(v,1):
                ws.cell(r,1,f"Cigno nero {i}"); ws.cell(r,2,json.dumps(s,ensure_ascii=False)); r+=1
        else:
            ws.cell(r,1,k); ws.cell(r,2,v); r+=1
    r+=1; ws.cell(r,1,"OUTPUT"); ws.cell(r,1).font=Font(bold=True); r+=1
    for c,col in enumerate(summary.columns,1): ws.cell(r,c,col).font=Font(bold=True)
    for _,rec in summary.iterrows():
        r+=1
        for c,col in enumerate(summary.columns,1): ws.cell(r,c,rec[col])

    for label,df in results.items():
        sh=wb.create_sheet(label[:31])
        for c,col in enumerate(df.columns,1): sh.cell(1,c,col).font=Font(bold=True)
        for rr,vals in enumerate(df.itertuples(index=False),2):
            for c,val in enumerate(vals,1): sh.cell(rr,c,val)

    ch=wb.create_sheet("Grafici"); ch["A1"]="Grafici"; ch["A1"].font=Font(bold=True,size=14)
    ages=results[next(iter(results))]["Età"].tolist()
    for metric, startrow, title in [
        ("Patrimonio monetario",3,"Patrimonio monetario"),
        ("Flusso netto dopo abitazione",60,"Flusso di cassa"),
        ("Patrimonio totale",117,"Patrimonio totale")
    ]:
        ch.cell(startrow,1,"Età")
        for j,label in enumerate(results.keys(),2): ch.cell(startrow,j,label)
        for i,age in enumerate(ages,startrow+1):
            ch.cell(i,1,age)
            for j,(label,df) in enumerate(results.items(),2):
                ch.cell(i,j,float(df.iloc[i-startrow-1][metric]))
        chart=LineChart(); chart.title=title; chart.x_axis.title="Età"; chart.y_axis.title="Euro"
        data=Reference(ch,min_col=2,max_col=1+len(results),min_row=startrow,max_row=startrow+len(ages))
        cats=Reference(ch,min_col=1,min_row=startrow+1,max_row=startrow+len(ages))
        chart.add_data(data,titles_from_data=True); chart.set_categories(cats); chart.height=10; chart.width=22
        ch.add_chart(chart,f"A{startrow+50}")
    wb.save(bio); bio.seek(0); return bio.getvalue()

# -----------------------------
# UI
# -----------------------------
st.title("LifePlan 360 — Alpha 4.2")
st.caption("Simulatore aperto 43/età scelta → 90 anni, con patrimonio monetario separato dall'immobiliare, previdenza, stress test e report.")

uploaded=st.file_uploader("Importa configurazione JSON",type=["json"])
u={}
if uploaded:
    try: u=json.load(uploaded); st.success("Configurazione importata")
    except Exception as e: st.error(str(e))

def gv(k,d): return u.get(k,d)

with st.sidebar:
    st.header("Anagrafica e obiettivi")
    age_start=st.number_input("Età iniziale",18,80,int(gv("age_start",43)))
    age_end=st.number_input("Età finale",max(int(age_start)+1,50),110,int(gv("age_end",90)))
    retire_age=st.number_input("Età uscita dal lavoro",int(age_start),90,int(gv("retire_age",59)))
    inps_age=st.number_input("Età minima pensionabile INPS",57,75,int(gv("inps_age",67)))
    target_final=st.number_input("Target patrimonio monetario finale (€)",0.0,value=float(gv("target_final",70000.0)),step=5000.0)

    st.header("Reddito e carriera")
    gross_salary=st.number_input("RAL iniziale (€)",0.0,value=float(gv("gross_salary",68000.0)),step=1000.0)
    salary_growth=st.number_input("Crescita RAL annua (%)",0.0,10.0,float(gv("salary_growth",0.02))*100,0.1,
                                  help="Guida: 1% prudente · 2% centrale · 3% dinamica")/100

    salary_mode=st.selectbox(
        "Modalità crescita carriera",
        ["Crescita costante","Scatti di carriera","Tabella personalizzata"],
        index=["Crescita costante","Scatti di carriera","Tabella personalizzata"].index(
            gv("salary_mode","Crescita costante")
        )
    )

    salary_steps=[]
    if salary_mode=="Scatti di carriera":
        st.caption("Gli scatti sostituiscono la RAL all'età indicata; poi riprende la crescita percentuale annua.")
        saved_steps=gv("salary_steps", [{},{},{}])
        for i in range(3):
            with st.expander(f"Scatto carriera {i+1}"):
                sv=saved_steps[i] if isinstance(saved_steps,list) and len(saved_steps)>i else {}
                active=st.checkbox("Attivo",value=bool(sv.get("active",False)),key=f"sal_step_on_{i}")
                step_age=st.number_input("Età",int(age_start),int(age_end),
                                         value=int(sv.get("age",min(int(age_end),int(age_start)+5*(i+1)))),
                                         key=f"sal_step_age_{i}")
                step_gross=st.number_input("Nuova RAL (€)",0.0,value=float(sv.get("gross",75000+10000*i)),
                                           step=1000.0,key=f"sal_step_gross_{i}")
                salary_steps.append({"active":active,"age":step_age,"gross":step_gross})

    salary_custom_points=[]
    if salary_mode=="Tabella personalizzata":
        st.caption("Inserisci fino a 6 punti età/RAL. Tra i punti il simulatore interpola la RAL.")
        saved_pts=gv("salary_custom_points", [])
        defaults=[]
        for i in range(6):
            sv=saved_pts[i] if isinstance(saved_pts,list) and len(saved_pts)>i else {}
            defaults.append({
                "Età": int(sv.get("age", min(int(age_end), int(age_start)+i*5))),
                "RAL": float(sv.get("gross", gross_salary*((1+salary_growth)**(i*5)))) if i < 4 else 0.0
            })
        salary_df=st.data_editor(pd.DataFrame(defaults),num_rows="fixed",use_container_width=True,key="salary_custom_editor")
        for _,r in salary_df.iterrows():
            if float(r["RAL"])>0:
                salary_custom_points.append({"age":int(r["Età"]),"gross":float(r["RAL"])})

    annual_bonus=st.number_input("Bonus annuo lordo iniziale (€)",0.0,value=float(gv("annual_bonus",0.0)),step=1000.0,
                                 help="MBO/premio medio annuo. Inserisci 0 se non previsto.")
    bonus_growth_with_salary=st.checkbox("Il bonus cresce con la stessa % della RAL",
                                         value=bool(gv("bonus_growth_with_salary",True)))

    st.header("Imposte sul lavoro")
    worker_contrib_rate=st.number_input("Contributi lavoratore in busta (%)",0.0,20.0,float(gv("worker_contrib_rate",0.0919))*100,0.1,help="Serve al netto di busta; distinto dall'aliquota INPS di computo.")/100
    regional_local_rate=st.number_input("Addizionali regionali/comunali stimate (%)",0.0,5.0,float(gv("regional_local_rate",0.02))*100,0.1)/100
    irpef1=st.number_input("IRPEF primo scaglione (%)",0.0,60.0,float(gv("irpef1",0.23))*100,0.5)/100
    irpef2=st.number_input("IRPEF secondo scaglione (%)",0.0,60.0,float(gv("irpef2",0.35))*100,0.5)/100
    irpef3=st.number_input("IRPEF terzo scaglione (%)",0.0,60.0,float(gv("irpef3",0.43))*100,0.5)/100

    st.header("INPS")
    inps_initial_montante=st.number_input("Montante contributivo INPS iniziale (€)",0.0,value=float(gv("inps_initial_montante",0.0)),step=5000.0)
    contrib_years_initial=st.number_input("Anni contributivi già maturati",0.0,60.0,float(gv("contrib_years_initial",15.0)),0.5)
    inps_computo_rate=st.number_input("Aliquota di computo INPS (%)",0.0,50.0,float(gv("inps_computo_rate",0.33))*100,0.5,help="Per dipendenti il default è 33% della retribuzione imponibile; non è la trattenuta personale in busta.")/100
    inps_revaluation=st.number_input("Rivalutazione montante INPS annua (%)",0.0,10.0,float(gv("inps_revaluation",0.015))*100,0.1)/100

    st.header("Fondo pensione / TFR")
    fund_initial=st.number_input("Fondo pensione iniziale (€)",0.0,value=float(gv("fund_initial",0.0)),step=5000.0)
    fund_years_initial=st.number_input("Anni già nel fondo pensione",0.0,60.0,float(gv("fund_years_initial",0.0)),0.5)
    rita_tax_rate=st.number_input("Tassazione media RITA stimata (%)",0.0,30.0,float(gv("rita_tax_rate",0.15))*100,0.5)/100

    st.header("Patrimonio e investimenti")
    initial_wealth=st.number_input("Patrimonio monetario iniziale (€)",0.0,value=float(gv("initial_wealth",260000.0)),step=5000.0)
    invest_share=st.slider("% patrimonio iniziale investito",0,100,int(gv("invest_share",0.80)*100))/100
    saving_invest_share=st.slider("% nuovi risparmi investiti",0,100,int(gv("saving_invest_share",0.80)*100))/100
    return_pre=st.number_input("Rendimento investimenti prima del de-risking (%)",-20.0,30.0,float(gv("return_pre",0.055))*100,0.1,help="Guida: 4% prudente · 5,5% centrale · 7% favorevole")/100
    derisk_age=st.number_input("Età riduzione rischio",int(age_start),100,int(gv("derisk_age",67)))
    return_post=st.number_input("Rendimento investimenti dopo de-risking (%)",-20.0,20.0,float(gv("return_post",0.02))*100,0.1,help="Guida: 1,5–3% profilo prudente")/100

    st.header("Spese / inflazione")
    living_monthly=st.number_input("Spese personali mensili iniziali (€)",0.0,value=float(gv("living_monthly",1600.0)),step=100.0,help="Escluse affitto, mutuo e manutenzioni.")
    inflation=st.number_input("Inflazione annua (%)",0.0,15.0,float(gv("inflation",0.02))*100,0.1,help="Guida: 1,5% bassa · 2% centrale · 3% stress · 4% severa")/100

    st.header("Affitto / proprietà")
    owns_home_initially=st.checkbox("Possiedo già la casa in cui vivo", value=bool(gv("owns_home_initially", False)))
    owned_home_value=st.number_input("Valore attuale casa già posseduta (€)",0.0,value=float(gv("owned_home_value",260000.0)),step=5000.0,disabled=not owns_home_initially)
    owned_home_ordinary_monthly=st.number_input("Manutenzione ordinaria mensile casa posseduta (€)",0.0,value=float(gv("owned_home_ordinary_monthly",200.0)),step=25.0,disabled=not owns_home_initially)
    owned_home_extra_annual=st.number_input("Manutenzione straordinaria media annua casa posseduta (€)",0.0,value=float(gv("owned_home_extra_annual",1800.0)),step=250.0,disabled=not owns_home_initially,help="Media annualizzata per lavori straordinari.")
    rent_monthly=st.number_input("Affitto mensile iniziale (€)",0.0,value=float(gv("rent_monthly",1300.0)),step=50.0)
    rent_growth=st.number_input("Crescita affitto annua (%)",0.0,15.0,float(gv("rent_growth",0.03))*100,0.1)/100
    house_price=st.number_input("Prezzo casa (€)",0.0,value=float(gv("house_price",260000.0)),step=5000.0)
    house_app=st.number_input("Rivalutazione casa annua (%)",-10.0,15.0,float(gv("house_app",0.015))*100,0.1,help="Guida Milano periferia: 0% molto prudente · 0,5% prudente · 1,5% centrale · 2–3% favorevole")/100
    mortgage_rate=st.number_input("Tasso mutuo (%)",0.0,20.0,float(gv("mortgage_rate",0.033))*100,0.1)/100
    mortgage_years=st.number_input("Durata mutuo (anni)",1,40,int(gv("mortgage_years",20)))
    buying_costs=st.number_input("Costi iniziali acquisto escluso anticipo (€)",0.0,value=float(gv("buying_costs",32000.0)),step=1000.0)
    owner_cost_monthly=st.number_input("Manutenzione/costi proprietà mensili (€)",0.0,value=float(gv("owner_cost_monthly",350.0)),step=25.0)

    st.header("Nuda proprietà")
    nuda_age=st.number_input("Età vendita nuda proprietà",50,100,int(gv("nuda_age",70)))
    nuda_pct=st.number_input("% valore casa incassato",0.0,100.0,float(gv("nuda_pct",0.60))*100,1.0)/100
    nuda_return=st.number_input("Rendimento capitale nuda proprietà (%)",-10.0,20.0,float(gv("nuda_return",0.015))*100,0.1)/100

    st.header("Downsizing")
    downsize_age=st.number_input("Età downsizing",50,100,int(gv("downsize_age",70)))
    smaller_house_today=st.number_input("Valore oggi casa più piccola (€)",0.0,value=float(gv("smaller_house_today",180000.0)),step=5000.0)
    downsize_costs=st.number_input("Costi compravendita downsizing (€)",0.0,value=float(gv("downsize_costs",25000.0)),step=1000.0)
    smaller_owner_cost_monthly=st.number_input("Costi mensili nuova casa (€)",0.0,value=float(gv("smaller_owner_cost_monthly",250.0)),step=25.0)

    st.header("Tre cigni neri")
    swans=[]
    for i in range(3):
        with st.expander(f"Cigno nero {i+1}"):
            active=st.checkbox("Attivo",value=bool(gv("swans",[{}, {}, {}])[i].get("active",False)) if isinstance(gv("swans",None),list) and len(gv("swans",[]))>i else False,key=f"a{i}")
            age=st.number_input("Età evento",int(age_start),int(age_end),int(gv("swans",[{}, {}, {}])[i].get("age",55+i*10)) if isinstance(gv("swans",None),list) and len(gv("swans",[]))>i else 55+i*10,key=f"age{i}")
            duration=st.number_input("Durata crisi (anni)",1,15,int(gv("swans",[{}, {}, {}])[i].get("duration",3)) if isinstance(gv("swans",None),list) and len(gv("swans",[]))>i else 3,key=f"d{i}")
            crash=st.number_input("Tonfo iniziale (%)",0.0,100.0,float(gv("swans",[{}, {}, {}])[i].get("crash",0.35))*100 if isinstance(gv("swans",None),list) and len(gv("swans",[]))>i else 35.0,1.0,key=f"c{i}")/100
            crisis_return=st.number_input("Rendimento annuo durante crisi (%)",-50.0,30.0,float(gv("swans",[{}, {}, {}])[i].get("crisis_return",0.0))*100 if isinstance(gv("swans",None),list) and len(gv("swans",[]))>i else 0.0,0.5,key=f"r{i}")/100
            swans.append({"active":active,"age":age,"duration":duration,"crash":crash,"crisis_return":crisis_return})

p=dict(
    age_start=age_start,age_end=age_end,retire_age=retire_age,inps_age=inps_age,target_final=target_final,
    gross_salary=gross_salary,salary_growth=salary_growth,salary_mode=salary_mode,
    salary_steps=salary_steps,salary_custom_points=salary_custom_points,
    annual_bonus=annual_bonus,bonus_growth_with_salary=bonus_growth_with_salary,
    worker_contrib_rate=worker_contrib_rate,
    regional_local_rate=regional_local_rate,irpef1=irpef1,irpef2=irpef2,irpef3=irpef3,
    inps_initial_montante=inps_initial_montante,contrib_years_initial=contrib_years_initial,
    inps_computo_rate=inps_computo_rate,inps_revaluation=inps_revaluation,
    fund_initial=fund_initial,fund_years_initial=fund_years_initial,rita_tax_rate=rita_tax_rate,
    initial_wealth=initial_wealth,invest_share=invest_share,saving_invest_share=saving_invest_share,
    return_pre=return_pre,derisk_age=derisk_age,return_post=return_post,
    living_monthly=living_monthly,inflation=inflation,
    owns_home_initially=owns_home_initially,owned_home_value=owned_home_value,owned_home_ordinary_monthly=owned_home_ordinary_monthly,owned_home_extra_annual=owned_home_extra_annual,
    rent_monthly=rent_monthly,rent_growth=rent_growth,house_price=house_price,house_app=house_app,
    mortgage_rate=mortgage_rate,mortgage_years=mortgage_years,buying_costs=buying_costs,
    owner_cost_monthly=owner_cost_monthly,nuda_age=nuda_age,nuda_pct=nuda_pct,nuda_return=nuda_return,
    downsize_age=downsize_age,smaller_house_today=smaller_house_today,downsize_costs=downsize_costs,
    smaller_owner_cost_monthly=smaller_owner_cost_monthly,swans=swans
)

if retire_age < age_start:
    st.error("L'età di uscita dal lavoro non può essere inferiore all'età iniziale.")

scenarios=[
    ("Affitto","rent",0.0,"keep"),
]
if owns_home_initially:
    scenarios.append(("Casa già posseduta","owned",0.0,"keep"))
scenarios += [
    ("Mutuo 20%","buy",0.20,"keep"),
    ("Mutuo 25%","buy",0.25,"keep"),
    ("Mutuo 30%","buy",0.30,"keep"),
    ("Mutuo 25% + Nuda","buy",0.25,"nuda"),
    ("Mutuo 25% + Downsizing","buy",0.25,"downsize"),
]
results={}; summaries=[]
for label,housing,dp,life in scenarios:
    df=simulate(p,housing,dp,life)
    results[label]=df
    summaries.append(summarize(df,p,label))
summary=pd.DataFrame(summaries)

st.subheader("Parametri stima reddito netto")
income_risk = pd.DataFrame([
    ["RAL iniziale", gross_salary, "Basso", "Input diretto dell'utente"],
    ["Modalità carriera", salary_mode, "Medio", "La carriera futura è intrinsecamente incerta"],
    ["Crescita RAL", salary_growth, "Medio", "Ipotesi di lungo periodo"],
    ["Bonus annuo iniziale", annual_bonus, "Medio-alto", "Può variare molto tra anni e aziende"],
    ["Contributi lavoratore in busta", worker_contrib_rate, "Medio", "Dipende da contratto e imponibile"],
    ["Addizionali regionali/comunali", regional_local_rate, "Medio-alto", "Dipendono da residenza e aliquote locali"],
    ["IRPEF primo scaglione", irpef1, "Basso", "Parametro normativo modificabile"],
    ["IRPEF secondo scaglione", irpef2, "Basso", "Parametro normativo modificabile"],
    ["IRPEF terzo scaglione", irpef3, "Basso", "Parametro normativo modificabile"],
    ["Stima reddito netto", "Calcolata", "Medio", "Non include tutte le detrazioni/deduzioni individuali"],
], columns=["Parametro","Valore","Rischio stima","Nota"])
st.dataframe(income_risk, use_container_width=True)

st.subheader("Specchietto finale scenari")
st.dataframe(summary.style.format({
    "Patrimonio monetario a 90":"€{:,.0f}",
    "Valore casa a 90":"€{:,.0f}",
    "Patrimonio totale a 90":"€{:,.0f}",
    "Patrimonio monetario minimo":"€{:,.0f}",
    "Pensione INPS netta annua":"€{:,.0f}"
}),use_container_width=True)

st.subheader("Grafici")
tab1,tab2,tab3,tab4=st.tabs(["Patrimonio monetario","Patrimonio totale","Flusso di cassa","Fondo pensione"])
with tab1:
    st.line_chart(pd.DataFrame({k:v.set_index("Età")["Patrimonio monetario"] for k,v in results.items()}))
with tab2:
    st.line_chart(pd.DataFrame({k:v.set_index("Età")["Patrimonio totale"] for k,v in results.items()}))
with tab3:
    st.line_chart(pd.DataFrame({k:v.set_index("Età")["Flusso netto dopo abitazione"] for k,v in results.items()}))
with tab4:
    st.line_chart(pd.DataFrame({k:v.set_index("Età")["Fondo pensione"] for k,v in results.items()}))

st.subheader("Registro input")
pretty = pd.DataFrame([(k, json.dumps(v,ensure_ascii=False) if isinstance(v,(list,dict)) else v) for k,v in p.items()],
                      columns=["Input","Valore"])
st.dataframe(pretty,use_container_width=True)

st.download_button("Esporta configurazione JSON",json.dumps(p,indent=2,ensure_ascii=False).encode("utf-8"),
                   "lifeplan_v4_config.json","application/json")

selected=st.selectbox("Dettaglio scenario",list(results.keys()))
st.dataframe(results[selected],use_container_width=True)
st.download_button("Scarica CSV scenario selezionato",results[selected].to_csv(index=False).encode("utf-8"),
                   f"{selected.replace(' ','_')}.csv","text/csv")

excel=excel_report(p,results,summary)
st.download_button("Scarica report Excel completo",excel,"LifePlan_v4_Report.xlsx",
                   "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

st.info("Nota: il calcolo fiscale, i coefficienti pensionistici e la RITA sono un modello di pianificazione, non una certificazione INPS/fiscale. Tutti i parametri principali sono esposti per consentire aggiornamenti futuri.")
