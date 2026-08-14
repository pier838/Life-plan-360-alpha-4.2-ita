
import json
import copy
import random
import math
from io import BytesIO

import numpy as np
import pandas as pd
import streamlit as st
from openpyxl import Workbook
from openpyxl.chart import LineChart, Reference
from openpyxl.styles import Font


st.set_page_config(page_title="LifePlan 360", page_icon="🏠", layout="wide")


# ============================================================
# Helpers
# ============================================================

def money(x):
    try:
        return f"€{float(x):,.0f}"
    except Exception:
        return str(x)


def progressive_tax(taxable, rates):
    tax = 0.0
    lower = 0.0
    for upper, rate in rates:
        part = max(0.0, min(taxable, upper) - lower)
        tax += part * rate
        lower = upper
        if taxable <= upper:
            break
    return max(0.0, tax)


def net_salary(gross, employee_contrib, local_tax, irpef1, irpef2, irpef3):
    contrib = gross * employee_contrib
    taxable = max(0.0, gross - contrib)
    tax = progressive_tax(
        taxable,
        [(28000.0, irpef1), (50000.0, irpef2), (10**12, irpef3)],
    )
    return max(0.0, gross - contrib - tax - taxable * local_tax)


def net_pension(gross, local_tax, irpef1, irpef2, irpef3):
    tax = progressive_tax(
        gross,
        [(28000.0, irpef1), (50000.0, irpef2), (10**12, irpef3)],
    )
    return max(0.0, gross - tax - gross * local_tax)


def annuity_payment(principal, annual_rate, years):
    if principal <= 0 or years <= 0:
        return 0.0
    r = annual_rate / 12
    n = int(years * 12)
    if abs(r) < 1e-12:
        return principal / n
    return principal * r / (1 - (1 + r) ** (-n))


def mortgage_balance(principal, annual_rate, years, elapsed_years):
    if principal <= 0 or elapsed_years >= years:
        return 0.0
    r = annual_rate / 12
    n = int(years * 12)
    m = int(max(0, elapsed_years * 12))
    if abs(r) < 1e-12:
        return max(0.0, principal * (1 - m / n))
    pmt = principal * r / (1 - (1 + r) ** (-n))
    return max(0.0, principal * (1 + r) ** m - pmt * (((1 + r) ** m - 1) / r))


INPS_COEFF = {
    57: 0.0420, 58: 0.0431, 59: 0.0442, 60: 0.0454, 61: 0.0466,
    62: 0.0480, 63: 0.0494, 64: 0.0509, 65: 0.0525, 66: 0.0542,
    67: 0.0561, 68: 0.0581, 69: 0.0602, 70: 0.0626, 71: 0.0651
}


def inps_coeff(age):
    if age <= 57:
        return INPS_COEFF[57]
    if age >= 71:
        return INPS_COEFF[71]
    return INPS_COEFF[int(age)]


def nuda_reference(age):
    pts = [
        (60, 0.40), (62, 0.45), (65, 0.50), (68, 0.55), (70, 0.60),
        (73, 0.65), (76, 0.70), (80, 0.75), (84, 0.80), (88, 0.85),
    ]
    if age <= pts[0][0]:
        return pts[0][1]
    for (a1, p1), (a2, p2) in zip(pts[:-1], pts[1:]):
        if a1 <= age <= a2:
            w = (age - a1) / (a2 - a1)
            return p1 + (p2 - p1) * w
    return pts[-1][1]


def validate_allocation(name, alloc):
    eq, bd, cash = alloc
    if eq + bd > 1.000001:
        st.error(
            f"{name}: Azionario + Governativi supera il 100% "
            f"({(eq+bd)*100:.1f}%). Riduci una delle due percentuali."
        )
        return False
    total = eq + bd + cash
    if abs(total - 1.0) > 0.001:
        st.error(f"{name}: allocazione non valida ({total*100:.1f}%).")
        return False
    return True


def career_gross(params, age):
    if age >= params["retire_age"]:
        return 0.0, 0.0, 0.0

    start = params["age_start"]
    base = params["gross_salary"]
    growth = params["salary_growth"]
    mode = params["salary_mode"]

    if mode == "Crescita costante":
        ral = base * ((1 + growth) ** (age - start))

    elif mode == "Scatti di carriera":
        ral = base
        steps = sorted(
            [x for x in params["salary_steps"] if x["active"]],
            key=lambda x: x["age"]
        )
        for a in range(start + 1, age + 1):
            ral *= (1 + growth)
            for step in steps:
                if step["age"] == a:
                    ral = step["gross"]

    else:
        pts = sorted(
            [(x["age"], x["gross"]) for x in params["salary_custom"] if x["gross"] > 0],
            key=lambda x: x[0]
        )
        if not pts:
            ral = base * ((1 + growth) ** (age - start))
        elif age <= pts[0][0]:
            ral = pts[0][1] / ((1 + growth) ** max(0, pts[0][0] - age))
        else:
            ral = None
            for (a1, g1), (a2, g2) in zip(pts[:-1], pts[1:]):
                if a1 <= age <= a2:
                    w = (age - a1) / (a2 - a1)
                    ral = g1 + (g2 - g1) * w
                    break
            if ral is None:
                la, lg = pts[-1]
                ral = lg * ((1 + growth) ** max(0, age - la))

    bonus = params["annual_bonus"]
    if params["bonus_growth"]:
        bonus *= (1 + growth) ** max(0, age - start)
    return ral, bonus, ral + bonus


def spending_band(params, age):
    if age < params["band1_end"]:
        return "Senior"
    if age < params["band2_end"]:
        return "Old"
    return "Older"


def annual_living_cost(params, age):
    band = spending_band(params, age)
    inflation_factor = (1 + params["inflation"]) ** (age - params["age_start"])
    values = {
        k: v * 12 * inflation_factor
        for k, v in params["expense_bands"][band].items()
    }
    return band, values, sum(values.values())


def allocation_for_age(params, age):
    # Timeline di mercato unica: questo determina solo l'esposizione corrente.
    # Non resetta rendimento, crisi o "anzianità" delle asset class.
    alloc=tuple(params["alloc_initial"])
    for stage in sorted([x for x in params.get("allocation_stages",[]) if x["active"]],key=lambda x:x["age"]):
        if age>=stage["age"]:
            alloc=tuple(stage["alloc"])
    return alloc


def rebalance(equity, bonds, cash, target, commission):
    total = equity + bonds + cash
    if total <= 0:
        return equity, bonds, cash, 0.0
    curr = np.array([equity, bonds, cash], dtype=float)
    targ = total * np.array(target, dtype=float)
    turnover = np.abs(targ - curr).sum() / 2
    fee = turnover * commission
    post = max(0.0, total - fee)
    return post * target[0], post * target[1], post * target[2], fee



def scenario_market_targets(params):
    """CAGR nominali target, già comprensivi di crisi e cigni neri."""
    mode = params.get("market_scenario", "Mediano")
    horizon = max(1, int(params["age_end"]) - int(params["age_start"]) + 1)

    if mode == "Manuale / Pro":
        return {
            "equity": float(params.get("pro_equity_cagr", 0.065)),
            "bonds": float(params.get("pro_bond_cagr", 0.035)),
        }

    base = {
        "Ottimista": {"equity": 0.080, "bonds": 0.040},
        "Mediano": {"equity": 0.065, "bonds": 0.035},
        "Pessimista": {"equity": 0.050, "bonds": 0.025},
    }[mode]

    curves = {
        "Ottimista": [(0,0.035),(5,0.045),(10,0.060),(15,0.070),(20,0.080)],
        "Mediano": [(0,0.020),(5,0.030),(10,0.045),(15,0.055),(20,0.065)],
        "Pessimista": [(0,0.000),(5,0.010),(10,0.020),(15,0.035),(20,0.050)],
    }[mode]

    def interp(points, x):
        if x >= points[-1][0]:
            return points[-1][1]
        for (x0,y0),(x1,y1) in zip(points[:-1], points[1:]):
            if x0 <= x <= x1:
                t=(x-x0)/(x1-x0)
                return y0+t*(y1-y0)
        return points[0][1]

    eq_target = interp(curves, min(horizon, 20))
    bond_target = base["bonds"]
    if horizon < 10:
        bond_target -= 0.0025
    elif horizon < 15:
        bond_target -= 0.0015
    elif horizon < 20:
        bond_target -= 0.0005

    return {"equity": eq_target, "bonds": bond_target}


def calibrated_normal_returns(params):
    """
    Calibra il rendimento annuo di fondo applicato OGNI anno.
    Crisi e recuperi sono moltiplicatori aggiuntivi.
    Il prodotto finale rispetta il CAGR nominale target dello scenario.
    """
    cal = deterministic_market_event_calendar(params)
    targets = scenario_market_targets(params)
    start = int(params["age_start"])
    end = int(params["age_end"])
    n = end - start + 1

    eq_event_factor = 1.0
    bd_event_factor = 1.0

    for e in cal["equity_crises"]:
        if not e.get("active", True):
            continue
        ev_start = int(e["age"])
        dur = max(1, int(e.get("duration", 3)))
        if start <= ev_start <= end:
            eq_event_factor *= (1 - float(e.get("drawdown", 0.25)))
        recovery_years = max(0, min(ev_start + dur - 1, end) - max(ev_start + 1, start) + 1)
        if recovery_years > 0:
            eq_event_factor *= (1 + float(e.get("recovery_return", 0.10))) ** recovery_years

    for e in cal["systemic_crises"]:
        if not e.get("active", True):
            continue
        ev_start = int(e["age"])
        dur = max(1, int(e.get("duration", 4)))
        if start <= ev_start <= end:
            eq_event_factor *= (1 - float(e.get("equity_drawdown", 0.35)))
            bd_event_factor *= (1 - float(e.get("bond_drawdown", 0.08)))
        recovery_years = max(0, min(ev_start + dur - 1, end) - max(ev_start + 1, start) + 1)
        if recovery_years > 0:
            eq_event_factor *= (1 + float(e.get("equity_recovery_return", 0.10))) ** recovery_years
            bd_event_factor *= (1 + float(e.get("bond_recovery_return", 0.04))) ** recovery_years

    eq_goal = (1 + targets["equity"]) ** n
    bd_goal = (1 + targets["bonds"]) ** n

    eq_base = (eq_goal / max(eq_event_factor, 1e-12)) ** (1 / n) - 1
    bd_base = (bd_goal / max(bd_event_factor, 1e-12)) ** (1 / n) - 1

    return {
        "equity_target": targets["equity"],
        "bond_target": targets["bonds"],
        "equity_normal": eq_base,
        "bond_normal": bd_base,
        "equity_event_factor": eq_event_factor,
        "bond_event_factor": bd_event_factor,
    }


def market_return_for_age(params, age):
    c = calibrated_normal_returns(params)
    return c["equity_normal"], c["bond_normal"]


def deterministic_market_event_calendar(params):
    """
    Timeline unica dei mercati per tutta la simulazione.

    IMPORTANTE:
    - il calendario dipende solo da scenario, sequenza e orizzonte complessivo;
    - NON riparte quando l'utente cambia asset allocation;
    - se a 75 anni resta il 10% azionario, una crisi a 75 anni colpisce solo quel 10%;
    - nuove obbligazioni acquistate con un ribilanciamento entrano nella timeline
      obbligazionaria già in corso, senza azzerare il conteggio degli anni.
    """
    mode = params.get("market_scenario", "Mediano")
    sequence = params.get("market_sequence", "Centrale")
    start = int(params["age_start"])
    end = int(params["age_end"])
    horizon = max(1, end - start + 1)

    if mode == "Manuale / Pro":
        return {
            "equity_crises": params.get("manual_equity_crises", []),
            "systemic_crises": params.get("manual_systemic_crises", []),
        }

    # Numero eventi: scala automaticamente con l'orizzonte.
    # Ordini di grandezza pensati per pianificazione, non come previsione puntuale.
    if mode == "Ottimista":
        eq_count = max(1, round(horizon / 13))
        sys_count = 0 if horizon < 35 else 1
        # Mix di crisi più leggere; grande crisi rara.
        equity_templates = [
            (0.20, 2, 0.13),
            (0.22, 2, 0.12),
            (0.24, 3, 0.11),
            (0.26, 3, 0.11),
            (0.30, 3, 0.12),
        ]
        systemic_templates = [
            (0.30, 0.05, 4, 0.13, 0.05),
        ]

    elif mode == "Pessimista":
        eq_count = max(2, round(horizon / 5.5))
        sys_count = 1 if horizon < 30 else 2 if horizon < 55 else 3
        equity_templates = [
            (0.22, 3, 0.08),
            (0.25, 3, 0.08),
            (0.28, 3, 0.085),
            (0.32, 4, 0.085),
            (0.36, 4, 0.09),
            (0.42, 5, 0.095),
            (0.50, 6, 0.10),
            (0.55, 6, 0.105),
        ]
        systemic_templates = [
            (0.42, 0.10, 5, 0.09, 0.035),
            (0.50, 0.15, 6, 0.095, 0.040),
            (0.55, 0.18, 6, 0.10, 0.045),
        ]

    else:  # Mediano
        eq_count = max(1, round(horizon / 7))
        sys_count = 0 if horizon < 25 else 1 if horizon < 50 else 2
        # Crisi eterogenee: più bear market moderati, poche crisi severe.
        equity_templates = [
            (0.20, 2, 0.12),
            (0.22, 2, 0.12),
            (0.24, 3, 0.11),
            (0.27, 3, 0.10),
            (0.30, 3, 0.10),
            (0.35, 4, 0.10),
            (0.45, 5, 0.11),
            (0.50, 5, 0.115),
        ]
        systemic_templates = [
            (0.36, 0.08, 4, 0.10, 0.045),
            (0.45, 0.12, 5, 0.105, 0.045),
        ]

    retire = int(params["retire_age"])

    def distributed_ages(count, lo, hi):
        if count <= 0:
            return []
        lo = int(lo)
        hi = int(max(lo, hi))
        if count == 1:
            return [int(round((lo + hi) / 2))]
        return [
            int(round(lo + (i + 1) / (count + 1) * (hi - lo)))
            for i in range(count)
        ]

    # La sequenza cambia il timing, non il numero di eventi né il CAGR target.
    if sequence == "Favorevole":
        eq_ages = distributed_ages(
            eq_count,
            start + 2,
            max(start + 3, min(retire - 2, end - 2))
        )
        sys_ages = distributed_ages(
            sys_count,
            start + 4,
            max(start + 5, min(retire - 3, end - 4))
        )
    elif sequence == "Sfavorevole":
        shock_start = max(start + 2, retire - 2)
        eq_ages = distributed_ages(eq_count, shock_start, end - 2)
        sys_ages = distributed_ages(
            sys_count,
            max(shock_start, retire),
            end - 3
        )
    else:
        eq_ages = distributed_ages(eq_count, start + 3, end - 3)
        sys_ages = distributed_ages(sys_count, start + 6, end - 5)

    # Evita che una crisi azionaria parta nello stesso identico anno del sistemico.
    sys_set = set(sys_ages)
    cleaned_eq = []
    for a in eq_ages:
        aa = a
        while aa in sys_set and aa < end - 1:
            aa += 1
        cleaned_eq.append(aa)

    def choose_templates(templates, count):
        """
        Se il numero eventi supera la lunghezza del template, riutilizza
        in modo deterministico il mix. Mantiene eterogeneità senza casualità.
        """
        if count <= 0:
            return []
        if count == 1:
            return [templates[len(templates)//2]]
        out = []
        for i in range(count):
            pos = round(i * (len(templates) - 1) / max(1, count - 1))
            out.append(templates[int(pos)])
        return out

    eq_specs = choose_templates(equity_templates, len(cleaned_eq))
    sys_specs = choose_templates(systemic_templates, len(sys_ages))

    equity_crises = []
    for age_evt, spec in zip(cleaned_eq, eq_specs):
        dd, duration, recovery = spec
        equity_crises.append({
            "active": True,
            "age": int(age_evt),
            "duration": int(duration),
            "drawdown": float(dd),
            "recovery_return": float(recovery),
        })

    systemic_crises = []
    for age_evt, spec in zip(sys_ages, sys_specs):
        eq_dd, bd_dd, duration, eq_rec, bd_rec = spec
        systemic_crises.append({
            "active": True,
            "age": int(age_evt),
            "duration": int(duration),
            "equity_drawdown": float(eq_dd),
            "bond_drawdown": float(bd_dd),
            "equity_recovery_return": float(eq_rec),
            "bond_recovery_return": float(bd_rec),
        })

    return {
        "equity_crises": equity_crises,
        "systemic_crises": systemic_crises,
    }



def apply_equity_only_crisis(equity, age, event):
    """Drawdown complessivo una sola volta, poi recupero."""
    if not event.get("active", True):
        return equity, ""
    start = int(event["age"])
    dur = max(1, int(event.get("duration", 3)))
    if not (start <= age < start + dur):
        return equity, ""
    if age == start:
        dd = float(event.get("drawdown", 0.25))
        return equity * (1 - dd), f"Crisi azionaria: drawdown complessivo -{dd*100:.1f}%"
    rec = float(event.get("recovery_return", 0.10))
    return equity * (1 + rec), f"Crisi azionaria: recupero +{rec*100:.1f}%"


def apply_systemic_crisis(equity, bonds, age, event):
    """Drawdown complessivo azioni/bond una sola volta, poi recuperi separati."""
    if not event.get("active", True):
        return equity, bonds, ""
    start = int(event["age"])
    dur = max(1, int(event.get("duration", 4)))
    if not (start <= age < start + dur):
        return equity, bonds, ""
    if age == start:
        edd = float(event.get("equity_drawdown", 0.35))
        bdd = float(event.get("bond_drawdown", 0.08))
        return (
            equity * (1 - edd),
            bonds * (1 - bdd),
            f"Cigno nero sistemico: drawdown complessivo azionario -{edd*100:.1f}% · obbligazionario -{bdd*100:.1f}%"
        )
    er = float(event.get("equity_recovery_return", 0.10))
    br = float(event.get("bond_recovery_return", 0.04))
    return (
        equity * (1 + er),
        bonds * (1 + br),
        f"Cigno nero sistemico: recupero azionario +{er*100:.1f}% · obbligazionario +{br*100:.1f}%"
    )


def crisis_active_labels(params,age):
    cal=deterministic_market_event_calendar(params); labels=[]
    for i,e in enumerate(cal["equity_crises"],1):
        if e.get("active",True) and e["age"]<=age<e["age"]+max(1,e.get("duration",3)): labels.append(f"Crisi azionaria {i}")
    for i,e in enumerate(cal["systemic_crises"],1):
        if e.get("active",True) and e["age"]<=age<e["age"]+max(1,e.get("duration",4)): labels.append(f"Cigno nero sistemico {i}")
    return labels


def unexpected_cost(params, age):
    total = 0.0
    labels = []
    for idx, event in enumerate(params["unexpected"], start=1):
        if event["active"] and age == event["age"]:
            amount = event["amount_today"] * (
                (1 + params["inflation"]) ** (age - params["age_start"])
            )
            total += amount
            labels.append(event["label"] or f"Imprevisto {idx}")
    return total, "; ".join(labels)


def late_care(params, age):
    if not params["late_care_active"] or age < params["late_care_age"]:
        return 0.0
    return params["late_care_monthly"] * 12 * (
        (1 + params["inflation"]) ** (age - params["age_start"])
    )


def fund_growth(params, fund, age):
    if not params.get("has_pension_fund", True):
        return 0.0
    eq, bd, ca = params.get("fund_alloc", [0.50,0.45,0.05])
    eq_r, bd_r = market_return_for_age(params, age)
    weighted = (
        eq * (eq_r - params["equity_cost"])
        + bd * (bd_r - params["bond_cost"])
        + ca * params["cash_return"]
    )
    return fund * (1 + weighted)


def apply_pension_market_shocks(params, fund, rita_pool, age):
    """
    Applica al fondo pensione e alla RITA gli stessi shock di mercato della
    timeline principale, pesati per l'asset allocation specifica del fondo.
    Il rendimento di fondo annuale viene applicato separatamente da fund_growth().
    """
    if not params.get("has_pension_fund", True):
        return 0.0, 0.0, []

    feq, fbd, fca = params.get("fund_alloc", [0.50, 0.45, 0.05])
    cal = deterministic_market_event_calendar(params)
    notes = []

    def apply_factor(value, factor):
        return max(0.0, value * max(0.0, factor))

    for crisis in cal["equity_crises"]:
        if not crisis.get("active", True):
            continue
        start = int(crisis["age"])
        dur = max(1, int(crisis.get("duration", 3)))
        if start <= age < start + dur:
            if age == start:
                dd = float(crisis.get("drawdown", 0.25))
                factor = feq * (1 - dd) + fbd + fca
                label = f"crisi azionaria fondo: quota azionaria -{dd*100:.1f}%"
            else:
                rec = float(crisis.get("recovery_return", 0.10))
                factor = feq * (1 + rec) + fbd + fca
                label = f"recupero crisi fondo: quota azionaria +{rec*100:.1f}%"
            fund = apply_factor(fund, factor)
            rita_pool = apply_factor(rita_pool, factor)
            notes.append(label)

    for crisis in cal["systemic_crises"]:
        if not crisis.get("active", True):
            continue
        start = int(crisis["age"])
        dur = max(1, int(crisis.get("duration", 4)))
        if start <= age < start + dur:
            if age == start:
                edd = float(crisis.get("equity_drawdown", 0.35))
                bdd = float(crisis.get("bond_drawdown", 0.08))
                factor = feq * (1 - edd) + fbd * (1 - bdd) + fca
                label = (
                    f"cigno nero fondo: azionario -{edd*100:.1f}% · "
                    f"obbligazionario -{bdd*100:.1f}%"
                )
            else:
                er = float(crisis.get("equity_recovery_return", 0.10))
                br = float(crisis.get("bond_recovery_return", 0.04))
                factor = feq * (1 + er) + fbd * (1 + br) + fca
                label = (
                    f"recupero cigno nero fondo: azionario +{er*100:.1f}% · "
                    f"obbligazionario +{br*100:.1f}%"
                )
            fund = apply_factor(fund, factor)
            rita_pool = apply_factor(rita_pool, factor)
            notes.append(label)

    return fund, rita_pool, notes


def market_cagr_audit(params):
    """
    Verifica indipendente del CAGR lordo della timeline di mercato su €1 iniziale.
    Serve a controllare che rendimento di fondo + shock = CAGR target.
    """
    eq = 1.0
    bd = 1.0
    start = int(params["age_start"])
    end = int(params["age_end"])
    n = end - start + 1
    cal = deterministic_market_event_calendar(params)

    for age in range(start, end + 1):
        eq_r, bd_r = market_return_for_age(params, age)
        eq *= (1 + eq_r)
        bd *= (1 + bd_r)

        for crisis in cal["equity_crises"]:
            eq, _ = apply_equity_only_crisis(eq, age, crisis)

        for crisis in cal["systemic_crises"]:
            eq, bd, _ = apply_systemic_crisis(eq, bd, age, crisis)

    return {
        "equity_actual": eq ** (1 / n) - 1,
        "bond_actual": bd ** (1 / n) - 1,
        "equity_final_factor": eq,
        "bond_final_factor": bd,
    }


def lifetime_annuity(capital, age, params):
    if capital <= 0:
        return 0.0
    end_age = max(age + 1, params["annuity_life_expectancy"])
    years = end_age - age
    rate = params["annuity_return"]
    if abs(rate) < 1e-12:
        gross = capital / years
    else:
        gross = capital * rate / (1 - (1 + rate) ** (-years))
    return gross * (1 - params["fund_payout_tax"])


# ============================================================
# Simulation
# ============================================================

def simulate(params, scenario):
    age0 = params["age_start"]
    age_end = params["age_end"]

    eq0, bd0, ca0 = params["alloc_initial"]
    if params.get("already_investor",True):
        equity=params["initial_wealth"]*eq0
        bonds=params["initial_wealth"]*bd0
        cash=params["initial_wealth"]*ca0
    else:
        equity=0.0
        bonds=0.0
        cash=params["initial_wealth"]

    house = 0.0
    debt = 0.0
    mortgage_principal = 0.0
    mortgage_monthly = 0.0

    nuda_reserve = 0.0
    nuda_done = False
    downsize_done = False

    if scenario["housing"] == "owned":
        house = params["owned_home_value"]

    purchase_done = scenario["housing"] != "buy"
    purchase_age = age0 + int(params.get("purchase_delay_years", 0))
    purchase_shortfall_total = 0.0

    # For buy scenarios, purchase is executed inside the yearly loop.
    # With delay 0 it happens at age_start; with delay N there are N rental years first.

    inps_montante = params["inps_initial_montante"]
    contrib_years = params["contrib_years_initial"]

    fund = params["fund_initial"]
    fund_years = params["fund_years_initial"]

    rita_active = False
    rita_pool = 0.0
    fund_settled = False
    complementary_annuity = 0.0

    inps_locked = False
    inps_net_annual = 0.0
    inps_gross_annual = 0.0

    rows = []

    for age in range(age0, age_end + 1):
        y = age - age0
        events = []

        band, expense_detail, normal_living = annual_living_cost(params, age)
        care_cost = late_care(params, age)
        living = normal_living + care_cost

        ral, bonus, gross = career_gross(params, age)
        work_net = 0.0
        if gross > 0:
            work_net = net_salary(
                gross,
                params["worker_contrib_rate"],
                params["add_rate"],
                params["irpef1"],
                params["irpef2"],
                params["irpef3"],
            )

        if age < params["retire_age"]:
            inps_montante *= (1 + params["inps_revaluation"])
            inps_montante += gross * params["inps_computo_rate"]
            contrib_years += 1

        # Previdenza complementare: fondo e RITA residua seguono SEMPRE
        # il proprio profilo di investimento fino alla liquidazione finale.
        # Se non esiste un fondo integrativo, il TFR non viene conferito qui.
        if params.get("has_pension_fund", True) and not fund_settled:
            fund = fund_growth(params, fund, age)
            if rita_pool > 0:
                rita_pool = fund_growth(params, rita_pool, age)

            fund, rita_pool, pension_market_notes = apply_pension_market_shocks(
                params, fund, rita_pool, age
            )
            if pension_market_notes and (fund > 0 or rita_pool > 0):
                events.extend([f"FONDO/RITA: {x}" for x in pension_market_notes])

            if age < params["retire_age"]:
                tfr = gross / 13.5
                fund += tfr
                fund_years += 1
        elif not params.get("has_pension_fund", True):
            fund = 0.0
            rita_pool = 0.0

        if age >= params["inps_age"] and not inps_locked:
            gross_pension = inps_montante * inps_coeff(params["inps_age"])
            inps_gross_annual = gross_pension
            inps_net_annual = net_pension(
                gross_pension,
                params["add_rate"],
                params["irpef1"],
                params["irpef2"],
                params["irpef3"],
            )
            inps_locked = True
            events.append("Inizio pensione INPS")

        rita_net = 0.0
        rita_min_age = max(params["retire_age"], params["inps_age"] - 5)
        rita_ok = (
            params.get("has_pension_fund",True)
            and params["use_rita"]
            and age >= rita_min_age
            and age < params["inps_age"]
            and contrib_years >= 20
            and fund_years >= 5
        )

        if rita_ok and (fund > 0 or rita_pool > 0):
            if rita_active:
                active_crises = crisis_active_labels(params, age)
                if active_crises:
                    events.append(
                        "WARNING RITA: erogazione durante "
                        + ", ".join(active_crises)
                        + " — sequence risk sul montante previdenziale"
                    )
            if not rita_active:
                rita_active = True
                rita_pool = fund * params["rita_share"]
                fund -= rita_pool
                events.append("Inizio RITA")
            years_left = max(1, params["inps_age"] - age)
            gross_rita = rita_pool / years_left
            rita_net = gross_rita * (1 - params["rita_tax"])
            rita_pool = max(0.0, rita_pool - gross_rita)

        capital_from_fund = 0.0
        if age >= params["inps_age"] and not fund_settled:
            total_fund = fund + rita_pool
            fund = 0.0
            rita_pool = 0.0

            mode = params["fund_at_inps_mode"]
            if mode == "Rendita vitalizia":
                complementary_annuity = lifetime_annuity(total_fund, age, params)
            elif mode == "Capitale + rendita":
                capital_from_fund = total_fund * params["fund_capital_share"]
                complementary_annuity = lifetime_annuity(
                    total_fund - capital_from_fund, age, params
                )
            else:
                capital_from_fund = total_fund
                complementary_annuity = 0.0

            fund_settled = True
            if total_fund > 0:
                events.append("Prestazione fondo pensione")

            if capital_from_fund > 0:
                teq, tbd, tca = allocation_for_age(params, age)
                equity += capital_from_fund * teq
                bonds += capital_from_fund * tbd
                cash += capital_from_fund * tca

        inps_income = inps_net_annual if age >= params["inps_age"] else 0.0
        comp_income = complementary_annuity if age >= params["inps_age"] else 0.0
        total_income = work_net + rita_net + inps_income + comp_income

        # Deferred purchase: before purchase the scenario behaves as rent.
        # The house price is expressed in today's/start-year euros and appreciates
        # until the actual purchase age. Buying costs are inflation-adjusted.
        purchase_shortfall = 0.0
        if (
            scenario["housing"] == "buy"
            and not purchase_done
            and age >= purchase_age
        ):
            delay = age - age0
            actual_price = params["house_price"] * ((1 + params["house_app"]) ** delay)
            actual_buying_costs = params["buying_costs"] * ((1 + params["inflation"]) ** delay)
            down = actual_price * scenario["down_pct"]
            upfront_need = down + actual_buying_costs

            # Fund upfront costs from available financial assets:
            # cash -> bonds -> equity. Never allow negative asset balances.
            use = min(max(0.0, cash), upfront_need)
            cash -= use
            upfront_need -= use

            use = min(max(0.0, bonds), upfront_need)
            bonds -= use
            upfront_need -= use

            use = min(max(0.0, equity), upfront_need)
            equity -= use
            upfront_need -= use

            purchase_shortfall = max(0.0, upfront_need)
            purchase_shortfall_total += purchase_shortfall

            if purchase_shortfall <= 1e-9:
                house = actual_price
                mortgage_principal = actual_price - down
                debt = mortgage_principal
                mortgage_monthly = annuity_payment(
                    mortgage_principal,
                    params["mortgage_rate"],
                    params["mortgage_years"],
                )
                purchase_done = True
                events.append(
                    f"Acquisto casa a {age} anni: prezzo {money(actual_price)}"
                )
            else:
                events.append(
                    f"INSOLVENZA ACQUISTO: anticipo/costi non finanziabili "
                    f"({money(purchase_shortfall)})"
                )
                # Prevent repeated attempts every following year.
                purchase_done = True

        # Appreciate house only after it exists; skip appreciation in the same
        # year it is purchased because actual_price already includes it.
        just_bought = (
            scenario["housing"] == "buy"
            and age == purchase_age
            and house > 0
        )
        if house > 0 and y > 0 and not just_bought:
            house *= (1 + params["house_app"])

        if scenario["housing"] == "rent":
            housing_cost = params["rent_monthly"] * 12 * (
                (1 + params["rent_growth"]) ** y
            )

        elif scenario["housing"] == "owned":
            housing_cost = home_value * (
                params.get("owned_home_maintenance_pct", 0.010)
                + params.get("owned_home_extra_pct", 0.005)
            )
            if params.get("owned_home_financing")=="Con mutuo in corso":
                remaining=max(0,params["owned_mortgage_months_left"]-y*12)
                housing_cost += params["owned_mortgage_payment"]*min(12,remaining)
            debt=0.0

        else:
            # Buy scenario: rent only BEFORE the purchase. After a nuda-proprieta
            # sale the user retains the right to live in the home, so no rent is due.
            if nuda_done:
                housing_cost = params["owner_cost_monthly"] * 12 * (
                    (1 + params["inflation"]) ** y
                )
                debt = 0.0
            elif house <= 0:
                housing_cost = params["rent_monthly"] * 12 * (
                    (1 + params["rent_growth"]) ** y
                )
                debt = 0.0
            else:
                elapsed_since_purchase = max(0, age - purchase_age)
                debt = mortgage_balance(
                    mortgage_principal,
                    params["mortgage_rate"],
                    params["mortgage_years"],
                    elapsed_since_purchase,
                )
                mortgage_annual = (
                    mortgage_monthly * 12
                    if elapsed_since_purchase < params["mortgage_years"]
                    else 0.0
                )
                housing_cost = (
                    mortgage_annual
                    + house * (
                        params.get("owned_home_maintenance_pct", 0.010)
                        + params.get("owned_home_extra_pct", 0.005)
                    )
                )

        nuda_proceeds_year = 0.0

        if (
            scenario.get("lifecycle") == "nuda"
            and not nuda_done
            and age >= params["nuda_age"]
            and house > 0
        ):
            proceeds = house * params["nuda_pct"]
            nuda_proceeds_year = proceeds
            nuda_reserve += proceeds
            house = 0.0
            debt = 0.0
            nuda_done = True
            events.append(f"Vendita nuda proprietà: incasso {money(proceeds)}")

        if (
            scenario.get("lifecycle") == "downsize"
            and not downsize_done
            and age >= params["downsize_age"]
            and house > 0
        ):
            smaller = params["smaller_house_today"] * (
                (1 + params["house_app"]) ** y
            )
            released = max(0.0, house - smaller - params["downsize_costs"])
            teq, tbd, tca = allocation_for_age(params, age)
            equity += released * teq
            bonds += released * tbd
            cash += released * tca
            house = smaller
            debt = 0.0
            housing_cost = params["smaller_owner_cost_monthly"] * 12 * (
                (1 + params["inflation"]) ** y
            )
            downsize_done = True
            events.append("Downsizing")

        # Eventuale ingresso iniziale nel mercato
        if not params.get("already_investor",True) and age>=params["invest_start_age"]:
            target=tuple(params["alloc_initial"])
            if params["invest_entry_mode"]=="Immediato" and age==params["invest_start_age"]:
                equity,bonds,cash,_=rebalance(equity,bonds,cash,target,0.0)
                events.append("Ingresso iniziale nel mercato")
            elif params["invest_entry_mode"]=="Graduale":
                yrs=max(1,params["invest_entry_years"])
                k=age-params["invest_start_age"]+1
                if 1<=k<=yrs:
                    total=max(0.0,equity+bonds+cash)
                    progress=min(1.0,k/yrs)
                    deq=total*target[0]*progress
                    dbd=total*target[1]*progress
                    move=min(cash,max(0.0,deq-equity)); cash-=move; equity+=move
                    move=min(cash,max(0.0,dbd-bonds)); cash-=move; bonds+=move
                    events.append(f"Ingresso graduale: {progress*100:.0f}%")

        for stage in params.get("allocation_stages",[]):
            if stage["active"] and age==stage["age"]:
                equity,bonds,cash,fee=rebalance(equity,bonds,cash,tuple(stage["alloc"]),stage["commission"])
                events.append(f"{stage['name']}: riallocazione, costo {money(fee)}")

        market_eq_return, market_bd_return = market_return_for_age(params, age)
        eq_gain = equity * (market_eq_return - params["equity_cost"])
        bd_gain = bonds * (market_bd_return - params["bond_cost"])
        cash_gain = cash * params["cash_return"]

        equity += eq_gain * (1 - params["equity_tax"])
        bonds += bd_gain * (1 - params["bond_tax"])
        cash += cash_gain * (1 - params["cash_tax"])

        market_calendar=deterministic_market_event_calendar(params)

        for crisis in market_calendar["equity_crises"]:
            equity,note=apply_equity_only_crisis(equity,age,crisis)
            if note: events.append(note)

        for crisis in market_calendar["systemic_crises"]:
            equity,bonds,note=apply_systemic_crisis(equity,bonds,age,crisis)
            if note: events.append(note)

        nuda_reserve *= (1 + params["nuda_return"])

        one_off, one_off_note = unexpected_cost(params, age)
        if one_off_note:
            events.append(one_off_note)

        outflows = living + housing_cost + one_off
        cashflow = total_income - outflows

        teq, tbd, tca = allocation_for_age(params, age)

        uncovered_need = 0.0
        sold_bonds = 0.0
        sold_equity = 0.0
        equity_sale_during_swan = False

        if cashflow >= 0:
            equity += cashflow * teq
            bonds += cashflow * tbd
            cash += cashflow * tca
        else:
            need = -cashflow

            use = min(max(0.0, nuda_reserve), need)
            nuda_reserve = max(0.0, nuda_reserve - use)
            need -= use

            use = min(max(0.0, cash), need)
            cash = max(0.0, cash - use)
            need -= use

            if need > 1e-9:
                use = min(max(0.0, bonds), need)
                bonds = max(0.0, bonds - use)
                need -= use
                sold_bonds += use
                if use > 0:
                    events.append(f"WARNING: disinvestimento governativi {money(use)}")

            if need > 1e-9:
                use = min(max(0.0, equity), need)
                equity = max(0.0, equity - use)
                need -= use
                sold_equity += use
                if use > 0:
                    swans_now = crisis_active_labels(params, age)
                    if swans_now:
                        equity_sale_during_swan = True
                        events.append(
                            "WARNING GRAVE: vendita forzata di azionario durante "
                            + ", ".join(swans_now)
                            + f" ({money(use)}) — rischio di sequenza dei rendimenti"
                        )
                    else:
                        events.append(
                            f"WARNING: decumulo da azionario {money(use)}"
                        )

            if need > 1e-9:
                uncovered_need = need
                events.append(
                    f"INSOLVENZA: fabbisogno non coperto {money(uncovered_need)}"
                )

        # Never allow financial asset balances to go below zero.
        equity = max(0.0, equity)
        bonds = max(0.0, bonds)
        cash = max(0.0, cash)
        nuda_reserve = max(0.0, nuda_reserve)

        fund_total = fund + rita_pool
        monetary = equity + bonds + cash + nuda_reserve + fund_total
        total_wealth = monetary + house - debt

        structural_cashflow = total_income - outflows
        cashflow_self_sufficient = structural_cashflow >= -1e-9
        inps_coverage_ratio = (
            inps_income / outflows
            if age >= params["inps_age"] and outflows > 0
            else 0.0
        )

        rows.append({
            "Età": age,
            "Fascia spese": band,
            "RAL base": ral,
            "Bonus lordo": bonus,
            "RAL complessiva": gross,
            "Netto lavoro": work_net,
            "RITA netta": rita_net,
            "Pensione INPS lorda": inps_gross_annual if age >= params["inps_age"] else 0.0,
            "Pensione INPS netta": inps_income,
            "Pensione integrativa": comp_income,
            "Reddito totale": total_income,
            "Cibo": expense_detail["Cibo"],
            "Utenze": expense_detail["Utenze"],
            "Trasporti": expense_detail["Trasporti"],
            "Svago e viaggi": expense_detail["Svago e viaggi"],
            "Salute": expense_detail["Salute"],
            "Assistenza tarda età": care_cost,
            "Costo abitazione": housing_cost,
            "Imprevisti": one_off,
            "Fabbisogno non coperto acquisto casa": purchase_shortfall,
            "Flusso netto": cashflow,
            "Flusso strutturale prima del patrimonio": structural_cashflow,
            "Autosufficiente nei flussi": cashflow_self_sufficient,
            "Spese totali annue": outflows,
            "Copertura INPS spese": inps_coverage_ratio,
            "Vendita governativi per decumulo": sold_bonds,
            "Vendita azionario per decumulo": sold_equity,
            "Vendita azionario durante cigno nero": equity_sale_during_swan,
            "Fabbisogno non coperto": uncovered_need,
            "Azionario": equity,
            "Obbligazionario": bonds,
            "Liquidità": cash,
            "Incasso nuda proprietà nell'anno": nuda_proceeds_year,
            "Riserva nuda proprietà": nuda_reserve,
            "Fondo pensione residuo": fund_total,
            "Montante INPS": inps_montante,
            "Valore casa": house,
            "Debito residuo": debt,
            "Patrimonio finanziario": monetary,
            "Patrimonio totale": total_wealth,
            "Eventi": "; ".join(events),
        })

    return pd.DataFrame(rows)


def summarize(df, params, label):
    neg_cf = df[df["Flusso netto"] < 0]
    min_money = float(df["Patrimonio finanziario"].min())
    final_money = float(df.iloc[-1]["Patrimonio finanziario"])
    final_house = float(df.iloc[-1]["Valore casa"])
    final_total = float(df.iloc[-1]["Patrimonio totale"])

    score = 0
    if (df["Patrimonio finanziario"] < 0).any():
        score += 5
    if final_money < params["target_final"]:
        score += 3
    elif final_money < 2 * params["target_final"]:
        score += 2
    elif final_money < 4 * params["target_final"]:
        score += 1
    if min_money < params["target_floor"]:
        score += 2

    if score <= 1:
        risk = "Basso"
    elif score <= 3:
        risk = "Medio"
    elif score <= 5:
        risk = "Medio-alto"
    else:
        risk = "Alto"

    robustness = max(0, 100 - score * 12)

    bond_sell = df[df["Vendita governativi per decumulo"] > 0]
    eq_sell = df[df["Vendita azionario per decumulo"] > 0]
    eq_swan = df[df["Vendita azionario durante cigno nero"] == True]
    insolv = df[df["Fabbisogno non coperto"] > 0]
    purchase_insolv = df[df["Fabbisogno non coperto acquisto casa"] > 0]

    post_inps = df[df["Età"] >= params["inps_age"]]
    if not post_inps.empty:
        first_inps = post_inps.iloc[0]
        inps_gross_start = float(first_inps["Pensione INPS lorda"])
        inps_net_start = float(first_inps["Pensione INPS netta"])
        expenses_inps_start = float(first_inps["Spese totali annue"])
        inps_coverage_start = float(first_inps["Copertura INPS spese"])
        post_inps_requires_decumulation = bool(
            (post_inps["Flusso strutturale prima del patrimonio"] < -1e-9).any()
        )
    else:
        inps_gross_start = 0.0
        inps_net_start = 0.0
        expenses_inps_start = 0.0
        inps_coverage_start = 0.0
        post_inps_requires_decumulation = True

    target_final_ok = final_money >= params["target_final"]
    no_insolvency = insolv.empty and purchase_insolv.empty

    if not insolv.empty or not purchase_insolv.empty:
        risk = "Alto"
        robustness = min(robustness, 15)
    elif not eq_swan.empty and risk in ("Basso", "Medio"):
        risk = "Medio-alto"
        robustness = min(robustness, 55)

    return {
        "Scenario": label,
        "Patrimonio finanziario finale": final_money,
        "Valore casa finale": final_house,
        "Patrimonio totale finale": final_total,
        "Patrimonio finanziario minimo": min_money,
        "Inizio decumulo": "-" if neg_cf.empty else int(neg_cf.iloc[0]["Età"]),
        "Prima vendita governativi": "-" if bond_sell.empty else int(bond_sell.iloc[0]["Età"]),
        "Prima vendita azionario": "-" if eq_sell.empty else int(eq_sell.iloc[0]["Età"]),
        "Anni con vendita azionario": int((df["Vendita azionario per decumulo"] > 0).sum()),
        "Vendita azionario in cigno nero": "Sì" if not eq_swan.empty else "No",
        "Età esaurimento patrimonio": "-" if insolv.empty else int(insolv.iloc[0]["Età"]),
        "Fabbisogno non coperto cumulato": float(df["Fabbisogno non coperto"].sum()),
        "Fabbisogno acquisto non coperto": float(df["Fabbisogno non coperto acquisto casa"].sum()),
        "Pensione INPS lorda annua a decorrenza": inps_gross_start,
        "Pensione INPS netta annua a decorrenza": inps_net_start,
        "Pensione INPS netta mensile equivalente": inps_net_start / 12 if inps_net_start else 0.0,
        "Spese annue alla decorrenza INPS": expenses_inps_start,
        "Copertura INPS delle spese": inps_coverage_start,
        "Pensione integrativa annua": float(df.iloc[-1]["Pensione integrativa"]),
        "Vincolo patrimoniale finale": "Rispettato" if target_final_ok else "Non rispettato",
        "Flussi dopo INPS": "Autosufficienti" if not post_inps_requires_decumulation else "Richiedono decumulo",
        "Rischio": risk,
        "Robustezza /100": robustness,
        "Sostenibilità complessiva": (
            "Pienamente sostenibile"
            if no_insolvency and target_final_ok
            else (
                "Flussi sostenibili, vincolo patrimoniale non raggiunto"
                if no_insolvency and not post_inps_requires_decumulation and not target_final_ok
                else (
                    "Sostenibile tramite decumulo, vincolo finale non raggiunto"
                    if no_insolvency and post_inps_requires_decumulation and not target_final_ok
                    else (
                        "Sostenibile tramite decumulo"
                        if no_insolvency and post_inps_requires_decumulation and target_final_ok
                        else "Non sostenibile / insolvenza"
                    )
                )
            )
        ),
    }


def build_excel(params, results, summary):
    bio = BytesIO()
    wb = Workbook()

    ws = wb.active
    ws.title = "Riepilogo"
    ws["A1"] = "LifePlan 360 — Input e output"
    ws["A1"].font = Font(bold=True, size=16)

    r = 3
    ws.cell(r, 1, "INPUT")
    ws.cell(r, 1).font = Font(bold=True)
    r += 1

    for key, value in params.items():
        ws.cell(r, 1, key)
        if isinstance(value, (dict, list, tuple)):
            value = json.dumps(value, ensure_ascii=False)
        ws.cell(r, 2, value)
        r += 1

    r += 1
    ws.cell(r, 1, "OUTPUT SCENARI")
    ws.cell(r, 1).font = Font(bold=True)
    r += 1

    for c, col in enumerate(summary.columns, start=1):
        ws.cell(r, c, col).font = Font(bold=True)

    for _, rec in summary.iterrows():
        r += 1
        for c, col in enumerate(summary.columns, start=1):
            val = rec[col]
            if hasattr(val, "item"):
                try:
                    val = val.item()
                except Exception:
                    pass
            ws.cell(r, c, val)

    for label, df in results.items():
        sh = wb.create_sheet(label[:31])
        for c, col in enumerate(df.columns, start=1):
            sh.cell(1, c, col).font = Font(bold=True)
        for rr, row in enumerate(df.itertuples(index=False), start=2):
            for c, val in enumerate(row, start=1):
                sh.cell(rr, c, val)

    ch = wb.create_sheet("Grafici")
    ch["A1"] = "Grafici comparativi"
    ch["A1"].font = Font(bold=True, size=16)

    ages = results[next(iter(results))]["Età"].tolist()
    configs = [
        ("Patrimonio finanziario", 3, "Patrimonio finanziario"),
        ("Patrimonio totale", 60, "Patrimonio totale"),
        ("Flusso netto", 117, "Flusso netto annuale"),
    ]

    for metric, start, title in configs:
        ch.cell(start, 1, "Età")
        for j, label in enumerate(results.keys(), start=2):
            ch.cell(start, j, label)

        for i, age in enumerate(ages, start=start + 1):
            ch.cell(i, 1, age)
            for j, (label, df) in enumerate(results.items(), start=2):
                idx = i - start - 1
                ch.cell(i, j, float(df.iloc[idx][metric]))

        chart = LineChart()
        chart.title = title
        chart.x_axis.title = "Età"
        chart.y_axis.title = "Euro"
        data = Reference(
            ch,
            min_col=2,
            max_col=1 + len(results),
            min_row=start,
            max_row=start + len(ages),
        )
        cats = Reference(
            ch,
            min_col=1,
            min_row=start + 1,
            max_row=start + len(ages),
        )
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        chart.width = 22
        chart.height = 10
        ch.add_chart(chart, f"A{start+50}")

    wb.save(bio)
    bio.seek(0)
    return bio.getvalue()


# ============================================================
# Interface
# ============================================================

st.markdown("# 🏠 💰 🚗 ✈️ LifePlan 360")
st.caption(
    "Simulatore della vita finanziaria: reddito, casa, investimenti, pensione, "
    "spese, imprevisti e resilienza."
)

uploaded = st.file_uploader("Importa configurazione JSON", type=["json"])
U = {}
if uploaded is not None:
    try:
        U = json.load(uploaded)
        st.success("Configurazione importata.")
    except Exception as exc:
        st.error(f"Configurazione JSON non valida: {exc}")



def scenario_list_for_params(params):
    suffix = "" if params.get("purchase_delay_years",0)==0 else f" (dopo {params['purchase_delay_years']} anni in affitto)"
    if params.get("owns_home_initially",False):
        return [
            {"label":"Casa già posseduta","housing":"owned","down_pct":0.0,"lifecycle":"keep"},
            {"label":"Casa posseduta + Nuda","housing":"owned","down_pct":0.0,"lifecycle":"nuda"},
            {"label":"Casa posseduta + Downsizing","housing":"owned","down_pct":0.0,"lifecycle":"downsize"},
        ]
    return [
        {"label":"Affitto","housing":"rent","down_pct":0.0,"lifecycle":"keep"},
        {"label":"Mutuo 20%"+suffix,"housing":"buy","down_pct":0.20,"lifecycle":"keep"},
        {"label":"Mutuo 25%"+suffix,"housing":"buy","down_pct":0.25,"lifecycle":"keep"},
        {"label":"Mutuo 30%"+suffix,"housing":"buy","down_pct":0.30,"lifecycle":"keep"},
        {"label":"Mutuo 25% + Nuda"+suffix,"housing":"buy","down_pct":0.25,"lifecycle":"nuda"},
        {"label":"Mutuo 25% + Downsizing"+suffix,"housing":"buy","down_pct":0.25,"lifecycle":"downsize"},
    ]


def mc_draw_params(base, rng):
    """Una traiettoria esogena: decisioni personali ferme, futuro economico variabile."""
    p = copy.deepcopy(base)

    # Mercati: scenari e sequenze sono variabili esogene, non decisioni dell'ottimizzatore.
    p["market_scenario"] = rng.choices(
        ["Ottimista","Mediano","Pessimista"], weights=[0.25,0.50,0.25], k=1
    )[0]
    p["market_sequence"] = rng.choice(["Favorevole","Centrale","Sfavorevole"])

    # Inflazione: shock persistente per traiettoria; le spese per fascia restano quelle scelte dall'utente.
    base_inf=float(base.get("inflation",0.02))
    p["inflation"] = min(0.065, max(0.0025, rng.gauss(base_inf,0.009)))

    # Casa e affitto: variabili nominali correlate imperfettamente all'inflazione.
    p["rent_growth"] = min(0.08, max(0.0, 0.55*p["inflation"] + 0.45*float(base.get("rent_growth",0.025)) + rng.gauss(0,0.007)))
    p["house_app"] = min(0.07, max(-0.025, float(base.get("house_app",0.015)) + 0.35*(p["inflation"]-base_inf) + rng.gauss(0,0.012)))

    # Liquidità: modesta sensibilità al regime inflattivo/tassi.
    p["cash_return"] = min(0.055, max(0.0, float(base.get("cash_return",0.005)) + 0.30*(p["inflation"]-base_inf) + rng.gauss(0,0.004)))
    return p


def mc_evaluate(base, scenario, runs=120, seed=360):
    """Valuta una strategia su futuri esogeni diversi. Successo = niente insolvenza + vincolo finale."""
    rng=random.Random(int(seed))
    finals=[]
    failures=0
    insolvencies=0
    forced_crisis_sales=0
    for _ in range(int(runs)):
        p=mc_draw_params(base,rng)
        df=simulate(p,scenario)
        final=float(df.iloc[-1]["Patrimonio finanziario"])
        finals.append(final)
        insolvent=bool((df["Fabbisogno non coperto"]>0).any() or (df["Fabbisogno non coperto acquisto casa"]>0).any())
        if insolvent:
            insolvencies += 1
        if bool(df["Vendita azionario durante cigno nero"].any()):
            forced_crisis_sales += 1
        if insolvent or final < float(base["target_final"]):
            failures += 1
    a=np.array(finals,dtype=float)
    return {
        "success_rate":1-failures/max(1,int(runs)),
        "insolvency_rate":insolvencies/max(1,int(runs)),
        "crisis_sale_rate":forced_crisis_sales/max(1,int(runs)),
        "p10":float(np.percentile(a,10)),
        "p50":float(np.percentile(a,50)),
        "p90":float(np.percentile(a,90)),
    }


def scale_expenses(params, multiplier):
    p=copy.deepcopy(params)
    p["expense_bands"]={
        band:{cat:float(v)*multiplier for cat,v in vals.items()}
        for band,vals in params["expense_bands"].items()
    }
    if p.get("late_care_active",False):
        p["late_care_monthly"]=float(params.get("late_care_monthly",0))*multiplier
    return p


def portfolio_risk_score(params):
    """Esposizione azionaria media ponderata per anni: più bassa = portafoglio più prudente."""
    vals=[]
    for age in range(int(params["age_start"]),int(params["age_end"])+1):
        vals.append(allocation_for_age(params,age)[0])
    return float(np.mean(vals)) if vals else 1.0


def portfolio_candidates(base, retire_age):
    """
    Glide path prudenti. L'ottimizzatore non conosce il futuro:
    sceglie una strategia prima delle traiettorie Monte Carlo.
    """
    start=int(base["age_start"]); end=int(base["age_end"])
    # Livelli azionari crescenti: si prova prima il rischio minore.
    levels=[0.10,0.15,0.20,0.25,0.30,0.35,0.40,0.45,0.50,0.55,0.60]
    candidates=[]
    for eq0 in levels:
        p=copy.deepcopy(base)
        p["retire_age"]=int(retire_age)
        # Manteniamo un cuscinetto liquido maggiore dopo il pensionamento.
        bd0=min(0.70, max(0.20, 0.85-eq0))
        ca0=max(0.05,1-eq0-bd0)
        p["alloc_initial"]=[eq0,bd0,ca0]

        ages=[
            min(end,max(start+1,int(retire_age))),
            min(end,max(start+2,int(retire_age)+7)),
            min(end,max(start+3,70)),
            min(end,max(start+4,80)),
        ]
        eqs=[max(0.05,eq0-0.05),max(0.05,eq0-0.12),max(0.05,eq0-0.22),max(0.02,eq0-0.30)]
        stages=[]
        for nm,ag,eq in zip(["Senior 1","Senior 2","Old","Older"],ages,eqs):
            cash=min(0.40,0.12 + 0.006*max(0,ag-retire_age))
            bond=max(0.0,1-eq-cash)
            stages.append({"name":nm,"active":True,"age":int(ag),"alloc":[eq,bond,cash],"commission":0.002})
        p["allocation_stages"]=stages
        candidates.append(p)
    return sorted(candidates,key=portfolio_risk_score)


def gv(key, default):
    return U.get(key, default)


with st.sidebar:
    st.header("1 · Anagrafica")
    age_start = st.number_input(
        "Età iniziale simulazione", 18, 85, int(gv("age_start", 43))
    )
    age_end = st.number_input(
        "Età finale simulazione",
        max(50, int(age_start) + 1),
        110,
        int(gv("age_end", 90)),
    )
    retire_age = st.number_input(
        "Età uscita dal lavoro",
        int(age_start),
        90,
        int(gv("retire_age", 59)),
    )
    inps_age = st.number_input(
        "Età pensione INPS", 57, 75, int(gv("inps_age", 67))
    )
    target_final = st.number_input(
        "Patrimonio finanziario minimo desiderato a fine simulazione (€)",
        0.0,
        value=float(gv("target_final", 70000.0)),
        step=5000.0,
    )
    target_floor = st.number_input(
        "Patrimonio finanziario minimo di sicurezza (€)",
        0.0,
        value=float(gv("target_floor", 30000.0)),
        step=5000.0,
    )

    st.header("2 · Reddito e carriera")
    gross_salary = st.number_input(
        "RAL iniziale (€)",
        0.0,
        value=float(gv("gross_salary", 68000.0)),
        step=1000.0,
    )
    salary_growth = st.number_input(
        "Crescita RAL annua (%)",
        0.0,
        10.0,
        float(gv("salary_growth", 0.02)) * 100,
        0.1,
        help="Guida: 1% prudente · 2% centrale · 3% dinamica",
    ) / 100

    salary_modes = ["Crescita costante", "Scatti di carriera", "Tabella personalizzata"]
    salary_mode = st.selectbox(
        "Modalità carriera",
        salary_modes,
        index=salary_modes.index(gv("salary_mode", "Crescita costante")),
    )

    salary_steps = []
    saved_steps = gv("salary_steps", [])
    if salary_mode == "Scatti di carriera":
        for i in range(3):
            saved = saved_steps[i] if isinstance(saved_steps, list) and len(saved_steps) > i else {}
            with st.expander(f"Scatto {i+1}"):
                active = st.checkbox(
                    "Attivo", bool(saved.get("active", False)), key=f"step_on_{i}"
                )
                step_age = st.number_input(
                    "Età",
                    int(age_start),
                    int(age_end),
                    int(saved.get("age", min(int(age_end), int(age_start) + 5 * (i + 1)))),
                    key=f"step_age_{i}",
                )
                step_gross = st.number_input(
                    "Nuova RAL (€)",
                    0.0,
                    value=float(saved.get("gross", 75000 + 10000 * i)),
                    step=1000.0,
                    key=f"step_gross_{i}",
                )
                salary_steps.append(
                    {"active": active, "age": int(step_age), "gross": step_gross}
                )

    salary_custom = []
    if salary_mode == "Tabella personalizzata":
        saved_custom = gv("salary_custom", [])
        rows = []
        for i in range(6):
            saved = (
                saved_custom[i]
                if isinstance(saved_custom, list) and len(saved_custom) > i
                else {}
            )
            rows.append({
                "Età": int(saved.get("age", min(int(age_end), int(age_start) + 5 * i))),
                "RAL": (
                    float(saved.get("gross", gross_salary * ((1 + salary_growth) ** (5 * i))))
                    if i < 4
                    else 0.0
                ),
            })
        edited = st.data_editor(
            pd.DataFrame(rows),
            num_rows="fixed",
            use_container_width=True,
            key="career_table",
        )
        for _, row in edited.iterrows():
            if float(row["RAL"]) > 0:
                salary_custom.append(
                    {"age": int(row["Età"]), "gross": float(row["RAL"])}
                )

    annual_bonus = st.number_input(
        "Bonus annuo lordo iniziale (€)",
        0.0,
        value=float(gv("annual_bonus", 0.0)),
        step=1000.0,
    )
    bonus_growth = st.checkbox(
        "Bonus cresce con la RAL", bool(gv("bonus_growth", True))
    )

    st.header("3 · Fiscalità lavoro")
    worker_contrib_rate = st.number_input(
        "Contributi lavoratore in busta (%)",
        0.0,
        20.0,
        float(gv("worker_contrib_rate", 0.0919)) * 100,
        0.1,
        help="Serve alla stima del netto; è distinto dall'aliquota di computo INPS.",
    ) / 100
    add_rate = st.number_input(
        "Addizionali regionali/comunali (%)",
        0.0,
        5.0,
        float(gv("add_rate", 0.02)) * 100,
        0.1,
    ) / 100
    irpef1 = st.number_input(
        "IRPEF primo scaglione (%)", 0.0, 60.0, float(gv("irpef1", 0.23)) * 100, 0.5
    ) / 100
    irpef2 = st.number_input(
        "IRPEF secondo scaglione (%)", 0.0, 60.0, float(gv("irpef2", 0.35)) * 100, 0.5
    ) / 100
    irpef3 = st.number_input(
        "IRPEF terzo scaglione (%)", 0.0, 60.0, float(gv("irpef3", 0.43)) * 100, 0.5
    ) / 100

    st.header("4 · Previdenza INPS")
    inps_initial_montante = st.number_input(
        "Montante INPS iniziale (€)",
        0.0,
        value=float(gv("inps_initial_montante", 0.0)),
        step=5000.0,
    )
    contrib_years_initial = st.number_input(
        "Anni contributivi già maturati",
        0.0,
        60.0,
        float(gv("contrib_years_initial", 15.0)),
        0.5,
    )
    inps_computo_rate = st.number_input(
        "Aliquota di computo INPS (%)",
        0.0,
        50.0,
        float(gv("inps_computo_rate", 0.33)) * 100,
        0.5,
        help="Default 33% per dipendente; non coincide con la sola trattenuta in busta.",
    ) / 100
    inps_revaluation = st.number_input(
        "Rivalutazione montante INPS annua (%)",
        0.0,
        10.0,
        float(gv("inps_revaluation", 0.015)) * 100,
        0.1,
    ) / 100

    st.header("5 · Fondo pensione / TFR")
    has_pension_fund = st.checkbox("Ho un fondo pensione integrativo", bool(gv("has_pension_fund", True)))
    fund_profile = st.selectbox(
        "Profilo fondo pensione",
        ["Garantito / molto prudente","Prudente","Bilanciato","Dinamico","Personalizzato"],
        disabled=not has_pension_fund,
    )
    fund_defaults={
        "Garantito / molto prudente":[0.05,0.90,0.05],
        "Prudente":[0.25,0.70,0.05],
        "Bilanciato":[0.50,0.45,0.05],
        "Dinamico":[0.75,0.20,0.05],
    }
    if fund_profile=="Personalizzato":
        fund_eq=st.number_input("Fondo: azionario (%)",0.0,100.0,50.0,1.0,disabled=not has_pension_fund)/100
        fund_bd=st.number_input("Fondo: obbligazionario (%)",0.0,100.0,45.0,1.0,disabled=not has_pension_fund)/100
        fund_ca=max(0.0,1-fund_eq-fund_bd)
    else:
        fund_eq,fund_bd,fund_ca=fund_defaults[fund_profile]
    st.caption(f"Fondo: {fund_eq*100:.0f}% azionario · {fund_bd*100:.0f}% obbligazionario · {fund_ca*100:.0f}% monetario")
    fund_initial = st.number_input(
        "Fondo pensione iniziale (€)",
        0.0,
        value=float(gv("fund_initial", 0.0)),
        step=5000.0,
    )
    fund_years_initial = st.number_input(
        "Anni già nel fondo pensione",
        0.0,
        60.0,
        float(gv("fund_years_initial", 0.0)),
        0.5,
    )
    use_rita = st.checkbox(
        "Utilizzare la RITA quando possibile",
        bool(gv("use_rita", True)) if has_pension_fund else False,
        disabled=not has_pension_fund,
    )
    rita_share = st.slider(
        "% fondo destinato alla RITA",
        0,
        100,
        int(gv("rita_share", 1.0) * 100),
    ) / 100
    rita_tax = st.number_input(
        "Tassazione RITA stimata (%)",
        0.0,
        30.0,
        float(gv("rita_tax", 0.15)) * 100,
        0.5,
    ) / 100

    fund_modes = [
        "Rendita vitalizia",
        "Capitale + rendita",
        "100% capitale (se consentito)",
    ]
    fund_at_inps_mode = st.selectbox(
        "Utilizzo fondo residuo alla pensione INPS",
        fund_modes,
        index=fund_modes.index(gv("fund_at_inps_mode", "Rendita vitalizia")),
    )
    fund_capital_share = st.slider(
        "% capitale nella modalità mista",
        0,
        100,
        int(gv("fund_capital_share", 0.50) * 100),
    ) / 100
    annuity_life_expectancy = st.number_input(
        "Età attesa per calcolo rendita vitalizia",
        70,
        110,
        int(gv("annuity_life_expectancy", 90)),
    )
    annuity_return = st.number_input(
        "Rendimento tecnico rendita (%)",
        0.0,
        10.0,
        float(gv("annuity_return", 0.015)) * 100,
        0.1,
    ) / 100
    fund_payout_tax = st.number_input(
        "Tassazione media rendita fondo (%)",
        0.0,
        30.0,
        float(gv("fund_payout_tax", 0.15)) * 100,
        0.5,
    ) / 100

    st.header("6 · Investimenti")
    initial_wealth = st.number_input(
        "Patrimonio finanziario iniziale (€)",
        0.0,
        value=float(gv("initial_wealth", 260000.0)),
        step=5000.0,
    )

    already_investor = st.checkbox(
        "Sei già investitore?", bool(gv("already_investor", True))
    )
    invest_start_age = int(age_start)
    invest_entry_mode = "Immediato"
    invest_entry_years = 0
    if not already_investor:
        invest_start_age = st.number_input(
            "Età inizio investimenti", int(age_start), int(age_end),
            int(gv("invest_start_age", int(age_start)))
        )
        invest_entry_mode = st.radio(
            "Ingresso nel mercato", ["Immediato", "Graduale"]
        )
        if invest_entry_mode == "Graduale":
            invest_entry_years = st.number_input(
                "Durata ingresso graduale (anni)", 1, 15,
                int(gv("invest_entry_years", 4))
            )

    st.subheader("Senior iniziale — sempre presente")
    st.caption(
        "Se gli stadi successivi sono disattivati, questo profilo resta valido fino a fine simulazione. "
        "Le crisi colpiscono solo la quota effettivamente investita nell'asset in quell'anno: "
        "ridurre l'azionario riduce automaticamente l'impatto delle crisi azionarie future."
    )
    saved_alloc = gv("alloc_initial", [0.60, 0.30, 0.10])
    ai_eq = st.number_input(
        "Azionario iniziale (%)", 0.0, 100.0,
        float(saved_alloc[0] * 100), 1.0
    ) / 100
    ai_bd = st.number_input(
        "Governativi iniziali (%)", 0.0, 100.0,
        float(saved_alloc[1] * 100), 1.0
    ) / 100
    ai_ca = max(0.0, 1.0 - ai_eq - ai_bd)
    st.caption(f"Liquidità automatica: {ai_ca*100:.1f}%")

    equity_return = st.number_input(
        "Rendimento ETF azionario legacy (%) — ignorato dal deterministico automatico",
        -20.0,
        30.0,
        float(gv("equity_return", 0.065)) * 100,
        0.1,
        help="Guida: 5% prudente · 6,5% centrale · 8% favorevole",
    ) / 100
    equity_tax = st.number_input(
        "Tassazione rendimenti azionari (%)",
        0.0,
        40.0,
        float(gv("equity_tax", 0.26)) * 100,
        0.5,
    ) / 100
    equity_cost = st.number_input(
        "TER/costo annuo ETF azionario (%)",
        0.0,
        5.0,
        float(gv("equity_cost", 0.002)) * 100,
        0.05,
    ) / 100

    bond_return = st.number_input(
        "Rendimento governativi legacy (%) — ignorato dal deterministico automatico",
        -10.0,
        20.0,
        float(gv("bond_return", 0.03)) * 100,
        0.1,
        help="Guida: 2% prudente · 3% centrale · 4% favorevole",
    ) / 100
    bond_tax = st.number_input(
        "Tassazione governativi (%)",
        0.0,
        30.0,
        float(gv("bond_tax", 0.125)) * 100,
        0.5,
    ) / 100
    bond_cost = st.number_input(
        "Costo annuo obbligazionario (%)",
        0.0,
        5.0,
        float(gv("bond_cost", 0.001)) * 100,
        0.05,
    ) / 100

    cash_return = st.number_input(
        "Rendimento nominale liquidità (%)",
        -5.0,
        10.0,
        float(gv("cash_return", 0.005)) * 100,
        0.1,
    ) / 100
    cash_tax = st.number_input(
        "Tassazione rendimento liquidità (%)",
        0.0,
        40.0,
        float(gv("cash_tax", 0.26)) * 100,
        0.5,
    ) / 100

    stage_guides = [
        ("Senior 1", 55, [0.55, 0.35, 0.10]),
        ("Senior 2", 62, [0.40, 0.45, 0.15]),
        ("Old", 70, [0.20, 0.55, 0.25]),
        ("Older", 80, [0.10, 0.50, 0.40]),
    ]
    allocation_stages = []
    for i, (nm, ag, guide) in enumerate(stage_guides):
        with st.expander(nm):
            active = st.checkbox(f"Attiva {nm}", False, key=f"st_on_{i}")
            age_stage = st.number_input(
                f"Età {nm}", int(age_start), int(age_end),
                min(int(age_end), max(int(age_start), ag)),
                key=f"st_age_{i}", disabled=not active
            )
            eq_stage = st.number_input(
                f"{nm}: azionario (%)",0.0,100.0,guide[0]*100,1.0,
                key=f"st_eq_{i}",disabled=not active)/100
            bd_stage = st.number_input(
                f"{nm}: obbligazionario (%)",0.0,100.0,guide[1]*100,1.0,
                key=f"st_bd_{i}",disabled=not active)/100
            ca_stage=max(0.0,1-eq_stage-bd_stage)
            st.caption(f"Liquidità automatica {ca_stage*100:.0f}% · guida {guide[0]*100:.0f}/{guide[1]*100:.0f}/{guide[2]*100:.0f}")
            comm=st.number_input(
                f"Costo riallocazione {nm} (%)",0.0,5.0,0.20,0.05,
                key=f"st_fee_{i}",disabled=not active)/100
            allocation_stages.append(
                {"name":nm,"active":active,"age":int(age_stage),
                 "alloc":[eq_stage,bd_stage,ca_stage],"commission":comm}
            )

    st.header("7 · Inflazione e spese")
    st.caption("Convenzione LifePlan: le spese delle fasce Senior/Old/Older sono inserite in euro di oggi. "
               "La simulazione le trasforma in euro nominali futuri applicando l’inflazione. "
               "I rendimenti finanziari sono anch’essi nominali: l’inflazione NON viene sottratta una seconda volta dai rendimenti.")
    inflation = st.number_input(
        "Inflazione annua (%)",
        0.0,
        15.0,
        float(gv("inflation", 0.02)) * 100,
        0.1,
        help="Guida: 1,5% bassa · 2% centrale · 3% stress · 4% severa",
    ) / 100

    default_band1 = min(max(int(age_start) + 1, 62), int(age_end) - 1)
    band1_end = st.number_input(
        "Fine fascia Senior",
        int(age_start) + 1,
        int(age_end) - 1,
        int(gv("band1_end", default_band1)),
    )

    default_band2 = min(max(int(band1_end) + 1, 70), int(age_end))
    band2_end = st.number_input(
        "Fine fascia Old",
        int(band1_end) + 1,
        int(age_end),
        int(gv("band2_end", default_band2)),
    )

    st.caption(
        f"Senior: {age_start}–{band1_end-1} · "
        f"Old: {band1_end}–{band2_end-1} · "
        f"Older: {band2_end}–{age_end}"
    )

    guide_defaults = {
        "Senior": {
            "Cibo": 450.0,
            "Utenze": 220.0,
            "Trasporti": 350.0,
            "Svago e viaggi": 500.0,
            "Salute": 120.0,
        },
        "Old": {
            "Cibo": 450.0,
            "Utenze": 230.0,
            "Trasporti": 300.0,
            "Svago e viaggi": 400.0,
            "Salute": 220.0,
        },
        "Older": {
            "Cibo": 450.0,
            "Utenze": 240.0,
            "Trasporti": 200.0,
            "Svago e viaggi": 200.0,
            "Salute": 400.0,
        },
    }

    saved_bands = gv("expense_bands", {})
    expense_bands = {}

    for band in ["Senior", "Old", "Older"]:
        vals = {}
        with st.expander(f"Spese {band}"):
            for cat in ["Cibo", "Utenze", "Trasporti", "Svago e viaggi", "Salute"]:
                default = float(
                    saved_bands.get(band, {}).get(cat, guide_defaults[band][cat])
                )
                vals[cat] = st.number_input(
                    f"{cat} €/mese",
                    0.0,
                    value=default,
                    step=25.0,
                    key=f"{band}_{cat}",
                )
        expense_bands[band] = vals

    st.subheader("Assistenza non autosufficienza")
    late_care_active = st.checkbox(
        "Considera assistenza in tarda età",
        bool(gv("late_care_active", False)),
    )
    late_care_age = st.number_input(
        "Età inizio assistenza",
        int(age_start),
        int(age_end),
        int(gv("late_care_age", min(85, int(age_end)))),
    )
    late_care_monthly = st.number_input(
        "Costo mensile in euro di oggi (€)",
        0.0,
        value=float(gv("late_care_monthly", 2300.0)),
        step=100.0,
        help="Voce guida per assistenza domiciliare/RSA; completamente modificabile.",
    )

    st.header("8 · Casa e affitto")
    owns_home_initially = st.checkbox(
        "Possiedo già la casa in cui vivo",
        bool(gv("owns_home_initially", False)),
    )

    # Default sempre definiti: l'interfaccia mostra solo i campi pertinenti,
    # ma il motore riceve sempre parametri validi.
    owned_home_value = 0.0
    owned_home_financing = "Già pagata / nessun mutuo"
    owned_mortgage_payment = 0.0
    owned_mortgage_months_left = 0

    house_price = 0.0
    purchase_delay_years = 0
    mortgage_rate = float(gv("mortgage_rate", 0.033))
    mortgage_years = int(gv("mortgage_years", 20))
    buying_costs = float(gv("buying_costs", 32000.0))

    # Affitto: serve sia nello scenario Affitto sia durante un eventuale
    # periodo di affitto precedente a un acquisto ritardato.
    rent_monthly = float(gv("rent_monthly", 1300.0))
    rent_growth = float(gv("rent_growth", 0.025))

    if owns_home_initially:
        owned_home_value = st.number_input(
            "Valore attuale della casa già posseduta (€)",
            0.0,
            value=float(gv("owned_home_value", 260000.0)),
            step=5000.0,
            help="Valore corrente dell'immobile in cui vivi già.",
        )
        owned_home_financing = st.radio(
            "La casa già posseduta è:",
            ["Già pagata / nessun mutuo", "Con mutuo in corso"],
            index=0 if gv("owned_home_financing", "Già pagata / nessun mutuo") == "Già pagata / nessun mutuo" else 1,
        )
        if owned_home_financing == "Con mutuo in corso":
            owned_mortgage_payment = st.number_input(
                "Rata mensile mutuo residuo (€)",
                0.0,
                value=float(gv("owned_mortgage_payment", 900.0)),
                step=50.0,
            )
            owned_mortgage_months_left = st.number_input(
                "Mensilità mancanti",
                0,
                600,
                int(gv("owned_mortgage_months_left", 120)),
            )
        st.caption(
            "Gli scenari di nuovo acquisto vengono esclusi automaticamente."
        )
        maintenance_reference_value = owned_home_value

    else:
        st.subheader("Affitto / acquisto")
        rent_monthly = st.number_input(
            "Affitto mensile iniziale (€)",
            0.0,
            value=float(gv("rent_monthly", 1300.0)),
            step=50.0,
            help="Usato nello scenario Affitto e nell'eventuale periodo precedente a un acquisto ritardato.",
        )
        rent_growth = st.number_input(
            "Crescita annua affitto (%)",
            0.0,
            15.0,
            float(gv("rent_growth", 0.025)) * 100,
            0.1,
            help="Guida: 1–2% prudente · 2–3% centrale · 3–4% pressione elevata · oltre 4% stress.",
        ) / 100
        rent_10y = rent_monthly * ((1 + rent_growth) ** 10)
        st.caption(
            f"Con questi input: €{rent_monthly:,.0f}/mese oggi → circa €{rent_10y:,.0f}/mese tra 10 anni."
        )

        house_price = st.number_input(
            "Valore/prezzo della casa da acquistare (€)",
            0.0,
            value=float(gv("house_price", 260000.0)),
            step=5000.0,
            help=(
                "Prezzo espresso nell'anno iniziale. Se l'acquisto è ritardato, "
                "il valore viene rivalutato fino all'anno effettivo."
            ),
        )
        purchase_delay_years = st.number_input(
            "Ritardo acquisto casa (anni)",
            0,
            max(0, int(age_end) - int(age_start)),
            int(gv("purchase_delay_years", 0)),
            help="0 = acquisto immediato; se >0, prima si vive in affitto.",
        )
        if purchase_delay_years == 0:
            st.caption("Scenario acquisto puro: la casa viene comprata nell'anno iniziale.")
        else:
            st.caption(
                f"Scenario misto: affitto per {purchase_delay_years} anni, "
                f"poi acquisto a {int(age_start)+int(purchase_delay_years)} anni."
            )

        mortgage_rate = st.number_input(
            "Tasso mutuo (%)",
            0.0,
            20.0,
            float(gv("mortgage_rate", 0.033)) * 100,
            0.1,
        ) / 100
        mortgage_years = st.number_input(
            "Durata mutuo (anni)",
            1,
            40,
            int(gv("mortgage_years", 20)),
        )
        buying_costs = st.number_input(
            "Costi acquisto escluso anticipo (€)",
            0.0,
            value=float(gv("buying_costs", 32000.0)),
            step=1000.0,
        )
        maintenance_reference_value = house_price

    house_app = st.number_input(
        "Rivalutazione casa annua (%)",
        -10.0,
        15.0,
        float(gv("house_app", 0.015)) * 100,
        0.1,
        help="Guida: 0% molto prudente · 0,5% prudente · 1,5% centrale · 2–3% favorevole.",
    ) / 100

    st.subheader("Costi di proprietà")
    owned_home_maintenance_pct = st.number_input(
        "Manutenzione ordinaria annua (% del valore casa)",
        0.0,
        5.0,
        float(gv("owned_home_maintenance_pct", 0.010)) * 100,
        0.1,
        help="Guida: 0,5% bassa · 1,0% centrale · 1,5% alta · 2,0%+ per immobili vecchi o impegnativi.",
    ) / 100
    owned_home_extra_pct = st.number_input(
        "Manutenzione straordinaria media annua (% del valore casa)",
        0.0,
        5.0,
        float(gv("owned_home_extra_pct", 0.005)) * 100,
        0.1,
        help="Guida come media annualizzata: circa 0,3–0,8% del valore della casa.",
    ) / 100

    ordinary_annual = maintenance_reference_value * owned_home_maintenance_pct
    extra_annual = maintenance_reference_value * owned_home_extra_pct
    st.caption(
        f"Con il valore casa impostato: ordinaria ≈ €{ordinary_annual:,.0f}/anno "
        f"(€{ordinary_annual/12:,.0f}/mese) · straordinaria ≈ €{extra_annual:,.0f}/anno "
        f"(€{extra_annual/12:,.0f}/mese)."
    )

    # Compatibilità con la logica esistente dei costi della casa acquistata:
    # usa l'equivalente mensile delle due percentuali all'anno iniziale.
    owner_cost_monthly = (ordinary_annual + extra_annual) / 12 if maintenance_reference_value > 0 else 0.0

    st.subheader("Nuda proprietà")
    nuda_age = st.number_input(
        "Età vendita nuda proprietà",
        50,
        100,
        int(gv("nuda_age", 70)),
    )
    nuda_guide = nuda_reference(nuda_age)
    st.caption(
        f"Guida indicativa alla percentuale sul valore pieno: circa {nuda_guide*100:.0f}%."
    )
    nuda_pct = st.number_input(
        "% valore casa incassato",
        0.0,
        100.0,
        float(gv("nuda_pct", nuda_guide)) * 100,
        1.0,
    ) / 100
    nuda_return = st.number_input(
        "Rendimento capitale nuda proprietà (%)",
        -10.0,
        20.0,
        float(gv("nuda_return", 0.015)) * 100,
        0.1,
    ) / 100

    st.subheader("Downsizing")
    downsize_age = st.number_input(
        "Età downsizing", 50, 100, int(gv("downsize_age", 70))
    )
    smaller_house_today = st.number_input(
        "Valore oggi casa più piccola (€)",
        0.0,
        value=float(gv("smaller_house_today", 180000.0)),
        step=5000.0,
    )
    downsize_costs = st.number_input(
        "Costi compravendita downsizing (€)",
        0.0,
        value=float(gv("downsize_costs", 25000.0)),
        step=1000.0,
    )
    smaller_owner_cost_monthly = st.number_input(
        "Costi nuova casa €/mese",
        0.0,
        value=float(gv("smaller_owner_cost_monthly", 250.0)),
        step=25.0,
    )

    st.header("9 · Scenario mercati")
    market_scenario=st.selectbox(
        "Scenario mercati deterministico",
        ["Ottimista","Mediano","Pessimista","Manuale / Pro"],
        index=["Ottimista","Mediano","Pessimista","Manuale / Pro"].index(gv("market_scenario","Mediano")),
        help="Gli scenari automatici sono fissi e riproducibili; Manuale / Pro consente di inserire direttamente gli eventi.",
    )
    market_sequence="Centrale"
    if market_scenario!="Manuale / Pro":
        market_sequence=st.selectbox(
            "Sequenza temporale delle crisi",
            ["Favorevole","Centrale","Sfavorevole"],
            index=["Favorevole","Centrale","Sfavorevole"].index(gv("market_sequence","Centrale")),
            help="Favorevole concentra maggiormente gli shock prima del pensionamento; Sfavorevole vicino o dopo l'uscita dal lavoro.",
        )
        st.caption("Stessi input + stessa sequenza = stesso risultato deterministico.")

    manual_equity_crises=[]; manual_systemic_crises=[]
    pro_equity_cagr = 0.065
    pro_bond_cagr = 0.035
    if market_scenario=="Manuale / Pro":
        st.subheader("Rendimenti target Manuale / Pro")
        pro_equity_cagr = st.number_input(
            "CAGR nominale target azionario (%) — crisi comprese",
            0.0, 20.0, float(gv("pro_equity_cagr",0.065))*100, 0.1
        ) / 100
        pro_bond_cagr = st.number_input(
            "CAGR nominale target obbligazionario (%) — shock compresi",
            0.0, 12.0, float(gv("pro_bond_cagr",0.035))*100, 0.1
        ) / 100
        st.caption(
            "Il software calcola il rendimento degli anni normali necessario "
            "a rispettare il CAGR target includendo le crisi inserite."
        )
        st.subheader("Crisi azionarie manuali")
        for i in range(3):
            with st.expander(f"Crisi azionaria {i+1}"):
                active=st.checkbox("Attiva",False,key=f"meq_on_{i}")
                age_evt=st.number_input("Età inizio",int(age_start),int(age_end),min(int(age_end),int(age_start)+8*(i+1)),key=f"meq_age_{i}",disabled=not active)
                dur=st.number_input("Durata totale (anni)",1,12,3,key=f"meq_dur_{i}",disabled=not active)
                dd=st.number_input("Drawdown massimo azionario (%)",0.0,90.0,25.0,1.0,key=f"meq_dd_{i}",disabled=not active)/100
                rec=st.number_input("Rendimento annuo recupero azionario (%)",-20.0,50.0,10.0,0.5,key=f"meq_rec_{i}",disabled=not active)/100
                manual_equity_crises.append({"active":active,"age":int(age_evt),"duration":int(dur),"drawdown":dd,"recovery_return":rec})

        st.subheader("Cigni neri sistemici manuali")
        for i in range(3):
            with st.expander(f"Cigno nero sistemico {i+1}"):
                active=st.checkbox("Attiva",False,key=f"msys_on_{i}")
                age_evt=st.number_input("Età inizio",int(age_start),int(age_end),min(int(age_end),int(age_start)+12*(i+1)),key=f"msys_age_{i}",disabled=not active)
                dur=st.number_input("Durata totale (anni)",1,15,4,key=f"msys_dur_{i}",disabled=not active)
                edd=st.number_input("Drawdown massimo azionario (%)",0.0,95.0,40.0,1.0,key=f"msys_edd_{i}",disabled=not active)/100
                bdd=st.number_input("Drawdown massimo obbligazionario (%)",0.0,60.0,10.0,1.0,key=f"msys_bdd_{i}",disabled=not active)/100
                er=st.number_input("Rendimento annuo recupero azionario (%)",-20.0,50.0,10.0,0.5,key=f"msys_er_{i}",disabled=not active)/100
                br=st.number_input("Rendimento annuo recupero obbligazionario (%)",-20.0,30.0,4.0,0.5,key=f"msys_br_{i}",disabled=not active)/100
                manual_systemic_crises.append({"active":active,"age":int(age_evt),"duration":int(dur),"equity_drawdown":edd,"bond_drawdown":bdd,"equity_recovery_return":er,"bond_recovery_return":br})

    st.header("10 · Imprevisti economici")
    unexpected = []
    saved_unexpected = gv("unexpected", [])
    for i in range(2):
        saved = (
            saved_unexpected[i]
            if isinstance(saved_unexpected, list) and len(saved_unexpected) > i
            else {}
        )
        with st.expander(f"Imprevisto {i+1}"):
            active = st.checkbox(
                "Attivo",
                bool(saved.get("active", False)),
                key=f"un_active_{i}",
            )
            event_age = st.number_input(
                "Età evento",
                int(age_start),
                int(age_end),
                int(saved.get("age", min(int(age_end), 65 + 10 * i))),
                key=f"un_age_{i}",
            )
            amount = st.number_input(
                "Perdita in euro di oggi (€)",
                0.0,
                value=float(saved.get("amount_today", 50000.0)),
                step=5000.0,
                key=f"un_amount_{i}",
            )
            label = st.text_input(
                "Descrizione",
                str(saved.get("label", "")),
                key=f"un_label_{i}",
            )
            unexpected.append({
                "active": active,
                "age": int(event_age),
                "amount_today": amount,
                "label": label,
            })


params = {
    "age_start": int(age_start),
    "age_end": int(age_end),
    "retire_age": int(retire_age),
    "inps_age": int(inps_age),
    "target_final": target_final,
    "target_floor": target_floor,
    "gross_salary": gross_salary,
    "salary_growth": salary_growth,
    "salary_mode": salary_mode,
    "salary_steps": salary_steps,
    "salary_custom": salary_custom,
    "annual_bonus": annual_bonus,
    "bonus_growth": bonus_growth,
    "worker_contrib_rate": worker_contrib_rate,
    "add_rate": add_rate,
    "irpef1": irpef1,
    "irpef2": irpef2,
    "irpef3": irpef3,
    "inps_initial_montante": inps_initial_montante,
    "contrib_years_initial": contrib_years_initial,
    "inps_computo_rate": inps_computo_rate,
    "inps_revaluation": inps_revaluation,
    "has_pension_fund": has_pension_fund,
    "fund_profile": fund_profile,
    "fund_alloc": [fund_eq,fund_bd,fund_ca],
    "fund_initial": fund_initial if has_pension_fund else 0.0,
    "fund_years_initial": fund_years_initial,
    "use_rita": use_rita,
    "rita_share": rita_share,
    "rita_tax": rita_tax,
    "fund_at_inps_mode": fund_at_inps_mode,
    "fund_capital_share": fund_capital_share,
    "annuity_life_expectancy": int(annuity_life_expectancy),
    "annuity_return": annuity_return,
    "fund_payout_tax": fund_payout_tax,
    "initial_wealth": initial_wealth,
    "already_investor": already_investor,
    "invest_start_age": int(invest_start_age),
    "invest_entry_mode": invest_entry_mode,
    "invest_entry_years": int(invest_entry_years),
    "alloc_initial": [ai_eq, ai_bd, ai_ca],
    "allocation_stages": allocation_stages,
    "equity_return": equity_return,
    "equity_tax": equity_tax,
    "equity_cost": equity_cost,
    "bond_return": bond_return,
    "bond_tax": bond_tax,
    "bond_cost": bond_cost,
    "cash_return": cash_return,
    "cash_tax": cash_tax,
    "inflation": inflation,
    "band1_end": int(band1_end),
    "band2_end": int(band2_end),
    "expense_bands": expense_bands,
    "late_care_active": late_care_active,
    "late_care_age": int(late_care_age),
    "late_care_monthly": late_care_monthly,
    "owns_home_initially": owns_home_initially,
    "owned_home_value": owned_home_value,
    "owned_home_financing": owned_home_financing,
    "owned_mortgage_payment": owned_mortgage_payment,
    "owned_mortgage_months_left": int(owned_mortgage_months_left),
    "owned_home_maintenance_pct": owned_home_maintenance_pct,
    "owned_home_extra_pct": owned_home_extra_pct,
    "rent_monthly": rent_monthly,
    "rent_growth": rent_growth,
    "house_price": house_price,
    "purchase_delay_years": int(purchase_delay_years),
    "house_app": house_app,
    "mortgage_rate": mortgage_rate,
    "mortgage_years": int(mortgage_years),
    "buying_costs": buying_costs,
    "owner_cost_monthly": owner_cost_monthly,
    "nuda_age": int(nuda_age),
    "nuda_pct": nuda_pct,
    "nuda_return": nuda_return,
    "downsize_age": int(downsize_age),
    "smaller_house_today": smaller_house_today,
    "downsize_costs": downsize_costs,
    "smaller_owner_cost_monthly": smaller_owner_cost_monthly,
    "market_scenario": market_scenario,
    "market_sequence": market_sequence,
    "pro_equity_cagr": pro_equity_cagr,
    "pro_bond_cagr": pro_bond_cagr,
    "manual_equity_crises": manual_equity_crises,
    "manual_systemic_crises": manual_systemic_crises,
    "unexpected": unexpected,
}


valid = True
valid &= validate_allocation("Senior iniziale", params["alloc_initial"])
for stage in params["allocation_stages"]:
    if stage["active"]:
        valid &= validate_allocation(stage["name"],stage["alloc"])
if params["has_pension_fund"]:
    valid &= validate_allocation("Fondo pensione",params["fund_alloc"])


st.subheader("Assunzioni scenario mercati")
market_calibration = calibrated_normal_returns(params)
market_audit = market_cagr_audit(params)
st.write(
    f"**Scenario:** {params['market_scenario']} · "
    f"**Sequenza:** {params.get('market_sequence','Manuale / Pro')} · "
    f"**Orizzonte:** {params['age_end']-params['age_start']+1} anni"
)
st.write(
    f"Azionario: **CAGR target {market_calibration['equity_target']*100:.2f}% nominale** "
    f"(crisi comprese) · rendimento annuo di fondo calibrato a "
    f"**{market_calibration['equity_normal']*100:.2f}%**."
)
st.write(
    f"Governativi: **CAGR target {market_calibration['bond_target']*100:.2f}% nominale** "
    f"(shock compresi) · rendimento annuo di fondo calibrato a "
    f"**{market_calibration['bond_normal']*100:.2f}%**."
)
st.write(
    f"Verifica matematica timeline: azionario **{market_audit['equity_actual']*100:.2f}% CAGR effettivo** · "
    f"governativi **{market_audit['bond_actual']*100:.2f}% CAGR effettivo**."
)
market_cagr_ok = (
    abs(market_audit["equity_actual"] - market_calibration["equity_target"]) < 0.0001
    and abs(market_audit["bond_actual"] - market_calibration["bond_target"]) < 0.0001
)
if not market_cagr_ok:
    st.error(
        "ERRORE CALIBRAZIONE MERCATI: il CAGR effettivo non coincide con il target. "
        "Non utilizzare questi risultati finché la calibrazione non è corretta."
    )
st.caption(
    "Tutti i rendimenti mostrati qui sono NOMINALI (prima dell'inflazione). "
    "I CAGR target includono già crisi e cigni neri: gli shock non vengono contati due volte. "
    "Le spese future sono anch'esse nominali perché vengono rivalutate con l'inflazione della simulazione. "
    "Quindi patrimonio e flussi di spesa sono confrontati sulla stessa base nominale. "
    "Per orizzonti brevi il target azionario automatico viene ridotto."
)
if market_calibration["equity_normal"] > 0.15 or market_calibration["bond_normal"] > 0.08:
    st.warning(
        "La combinazione CAGR/shock richiede un rendimento annuo di fondo molto elevato: "
        "verifica la plausibilità delle assunzioni."
    )

st.subheader("Calendario deterministico dei mercati")
cal_preview=deterministic_market_event_calendar(params)
preview_rows=[]
for i,e in enumerate(cal_preview["equity_crises"],1):
    if e.get("active",True): preview_rows.append({"Tipo":f"Crisi azionaria {i}","Età":e["age"],"Durata":e.get("duration",3),"Drawdown azionario":f"{e.get('drawdown',0)*100:.0f}%","Drawdown obbligazionario":"0%"})
for i,e in enumerate(cal_preview["systemic_crises"],1):
    if e.get("active",True): preview_rows.append({"Tipo":f"Cigno nero sistemico {i}","Età":e["age"],"Durata":e.get("duration",4),"Drawdown azionario":f"{e.get('equity_drawdown',0)*100:.0f}%","Drawdown obbligazionario":f"{e.get('bond_drawdown',0)*100:.0f}%"})
if preview_rows: st.dataframe(pd.DataFrame(preview_rows),use_container_width=True,hide_index=True)
else: st.caption("Nessun evento sistemico automatico previsto per questo orizzonte/scenario.")

st.subheader("Assunzioni principali e affidabilità")
market_diag = calibrated_normal_returns(params)
assumptions = pd.DataFrame([
    ["RAL iniziale", money(gross_salary), "Alta", "Dato inserito dall'utente"],
    ["Crescita RAL", f"{salary_growth*100:.1f}%", "Media", "Carriera futura incerta"],
    ["Stima netto", "Modello parametrico", "Media", "Non include ogni detrazione/deduzione personale"],
    ["Inflazione", f"{inflation*100:.1f}%", "Bassa-Media", "Usata per trasformare le spese in euro nominali futuri"],
    ["Convenzione monetaria", "Nominale", "Alta", "Rendimenti e flussi futuri sono entrambi nominali; nessuna doppia sottrazione dell’inflazione"],
    ["Rivalutazione casa", f"{house_app*100:.1f}%", "Bassa-Media", "Dipende da zona e immobile"],
    ["Pensione INPS", "Montante × coefficiente", "Media", "Stima di pianificazione"],
    ["Scenario mercati", params["market_scenario"], "Media", "Scenario deterministico selezionato"],
    ["CAGR azionario target", f"{market_diag['equity_target']*100:.2f}%", "Media", "Nominale, crisi comprese"],
    ["CAGR governativi target", f"{market_diag['bond_target']*100:.2f}%", "Media", "Nominale, shock compresi"],
    ["Rendimento azionario di fondo", f"{market_diag['equity_normal']*100:.2f}%", "Media", "Calibrato per rispettare il CAGR target"],
    ["Rendimento governativi di fondo", f"{market_diag['bond_normal']*100:.2f}%", "Media", "Calibrato per rispettare il CAGR target"],
], columns=["Parametro", "Valore", "Affidabilità", "Nota"])
st.dataframe(assumptions, use_container_width=True, hide_index=True)


st.markdown("---")
st.caption("La simulazione non viene ricalcolata mentre modifichi gli input: parte solo quando premi il pulsante.")
# Blocco di sicurezza: non eseguire scenari con calibrazione mercati incoerente.
if "market_cagr_ok" in globals():
    valid = valid and market_cagr_ok

run_simulation = st.button(
    "▶️ Esegui simulazione",
    type="primary",
    use_container_width=True,
    disabled=not valid,
)

if run_simulation:
    st.session_state["lifeplan_run_params"] = params

if "lifeplan_run_params" in st.session_state:
    if not run_simulation:
        st.info(
            "Hai modificato degli input? Premi **Esegui simulazione** per aggiornare i risultati. "
            "Qui sotto restano visibili i risultati dell'ultima esecuzione."
        )
    params = st.session_state["lifeplan_run_params"]

    scenarios = scenario_list_for_params(params)

    results = {}
    summaries = []

    for scenario in scenarios:
        df = simulate(params, scenario)
        results[scenario["label"]] = df
        summaries.append(summarize(df, params, scenario["label"]))

    summary = pd.DataFrame(summaries)

    st.subheader("Specchietto finale scenari")
    st.dataframe(
        summary.style.format({
            "Patrimonio finanziario finale": "€{:,.0f}",
            "Valore casa finale": "€{:,.0f}",
            "Patrimonio totale finale": "€{:,.0f}",
            "Patrimonio finanziario minimo": "€{:,.0f}",
            "Pensione INPS lorda annua a decorrenza": "€{:,.0f}",
            "Pensione INPS netta annua a decorrenza": "€{:,.0f}",
            "Pensione INPS netta mensile equivalente": "€{:,.0f}",
            "Spese annue alla decorrenza INPS": "€{:,.0f}",
            "Copertura INPS delle spese": "{:.1%}",
            "Pensione integrativa annua": "€{:,.0f}",
            "Fabbisogno non coperto cumulato": "€{:,.0f}",
            "Fabbisogno acquisto non coperto": "€{:,.0f}",
        }),
        use_container_width=True,
    )

    st.subheader("Pensione INPS e fattibilità dei flussi")
    st.caption(
        "Il giudizio sui flussi è separato dal vincolo patrimoniale finale. "
        "Dopo la decorrenza INPS, LifePlan verifica se pensione e altri redditi correnti "
        "coprono le spese dello scenario oppure se è necessario continuare a decumulare patrimonio."
    )

    cashflow_rows = []
    for scenario_name, df in results.items():
        post = df[df["Età"] >= params["inps_age"]]
        if not post.empty:
            r = post.iloc[0]
            summary_row = summary[summary["Scenario"] == scenario_name].iloc[0]
            cashflow_rows.append({
                "Scenario": scenario_name,
                "Età INPS": int(params["inps_age"]),
                "INPS lorda annua": float(r["Pensione INPS lorda"]),
                "INPS netta annua": float(r["Pensione INPS netta"]),
                "INPS netta/mese": float(r["Pensione INPS netta"]) / 12,
                "Spese annue a INPS": float(r["Spese totali annue"]),
                "Copertura INPS": float(r["Copertura INPS spese"]),
                "Flussi dopo INPS": summary_row["Flussi dopo INPS"],
                "Vincolo finale": summary_row["Vincolo patrimoniale finale"],
                "Esito": summary_row["Sostenibilità complessiva"],
            })

    if cashflow_rows:
        cashflow_df = pd.DataFrame(cashflow_rows)
        st.dataframe(
            cashflow_df.style.format({
                "INPS lorda annua": "€{:,.0f}",
                "INPS netta annua": "€{:,.0f}",
                "INPS netta/mese": "€{:,.0f}",
                "Spese annue a INPS": "€{:,.0f}",
                "Copertura INPS": "{:.1%}",
            }),
            use_container_width=True,
            hide_index=True,
        )

    tab1, tab2, tab3, tab4 = st.tabs([
        "Patrimonio finanziario",
        "Patrimonio totale",
        "Flusso netto",
        "Asset allocation",
    ])

    with tab1:
        st.line_chart(
            pd.DataFrame({
                k: v.set_index("Età")["Patrimonio finanziario"]
                for k, v in results.items()
            })
        )

    with tab2:
        st.line_chart(
            pd.DataFrame({
                k: v.set_index("Età")["Patrimonio totale"]
                for k, v in results.items()
            })
        )

    with tab3:
        st.line_chart(
            pd.DataFrame({
                k: v.set_index("Età")["Flusso netto"]
                for k, v in results.items()
            })
        )

    with tab4:
        asset_scenario = st.selectbox(
            "Scenario asset allocation",
            list(results.keys()),
            key="asset_scenario",
        )
        st.line_chart(
            results[asset_scenario]
            .set_index("Età")[["Azionario", "Obbligazionario", "Liquidità"]]
        )

    st.subheader("Timeline eventi")
    timeline_scenario = st.selectbox(
        "Scenario timeline",
        list(results.keys()),
        key="timeline_scenario",
    )
    timeline = results[timeline_scenario]
    timeline = timeline[timeline["Eventi"] != ""][["Età", "Eventi"]]
    st.dataframe(timeline, use_container_width=True, hide_index=True)

    st.subheader("Dettaglio annuale")
    detail_scenario = st.selectbox(
        "Scenario dettagliato",
        list(results.keys()),
        key="detail_scenario",
    )
    st.dataframe(results[detail_scenario], use_container_width=True)


    st.markdown("---")
    st.header("🧭 Ottimizzatori LifePlan")
    st.caption(
        "Gli ottimizzatori partono SOLO quando premi il relativo pulsante. "
        "Le spese per fascia restano espresse in euro di oggi. Nel Monte Carlo variano "
        "mercati, sequenza delle crisi, inflazione, crescita affitto, rivalutazione casa e rendimento della liquidità."
    )

    opt_scenario_label = st.selectbox(
        "Scenario abitativo da ottimizzare",
        list(results.keys()),
        key="optimizer_housing_scenario",
    )
    opt_scenario = next(x for x in scenarios if x["label"] == opt_scenario_label)
    mc_runs = st.select_slider(
        "Numero simulazioni Monte Carlo per test",
        options=[40,80,120,200,400],
        value=120,
        help="Per i primi test usa 80–120. 400 è più robusto ma può essere lento.",
        key="optimizer_mc_runs",
    )
    success_threshold = st.slider(
        "Soglia minima di successo Monte Carlo",
        0.70,0.99,0.90,0.01,
        help="Successo = nessuna insolvenza e patrimonio finanziario finale almeno pari al vincolo impostato.",
        key="optimizer_success_threshold",
    )

    with st.expander("1 · Ottimizza Pensione", expanded=True):
        st.write(
            "**Domanda:** con le spese e il portafoglio che hai impostato, qual è la prima età "
            "alla quale puoi smettere di lavorare rispettando la soglia Monte Carlo?"
        )
        min_ret_age=st.number_input(
            "Età minima da testare", int(params["age_start"])+1, int(params["inps_age"]),
            min(max(int(params["age_start"])+1,50),int(params["inps_age"])),
            key="opt_ret_min"
        )
        max_ret_age=st.number_input(
            "Età massima da testare", int(min_ret_age), int(params["inps_age"]),
            int(params["inps_age"]), key="opt_ret_max"
        )
        if st.button("🎯 Ottimizza età pensionamento", use_container_width=True):
            prog=st.progress(0)
            rows_opt=[]
            ages=list(range(int(min_ret_age),int(max_ret_age)+1))
            for j,age in enumerate(ages):
                pp=copy.deepcopy(params); pp["retire_age"]=age
                m=mc_evaluate(pp,opt_scenario,mc_runs,seed=360+age)
                rows_opt.append({"Età uscita":age,"Successo MC":m["success_rate"],
                                 "P10 finale":m["p10"],"P50 finale":m["p50"],
                                 "Insolvenza":m["insolvency_rate"],
                                 "Vendita azioni in cigno nero":m["crisis_sale_rate"],
                                 "Fattibile":m["success_rate"]>=success_threshold})
                prog.progress((j+1)/len(ages))
            odf=pd.DataFrame(rows_opt)
            feasible=odf[odf["Fattibile"]]
            if feasible.empty:
                st.error("Nessuna età testata raggiunge la soglia di sicurezza impostata.")
            else:
                best=int(feasible.iloc[0]["Età uscita"])
                st.success(f"Prima età fattibile: **{best} anni**.")
            st.dataframe(odf.style.format({"Successo MC":"{:.1%}","P10 finale":"€{:,.0f}",
                "P50 finale":"€{:,.0f}","Insolvenza":"{:.1%}",
                "Vendita azioni in cigno nero":"{:.1%}"}),use_container_width=True,hide_index=True)

    with st.expander("2 · Ottimizza Stile di Vita"):
        st.write(
            "**Domanda:** fissata l'età di uscita, qual è la massima spesa sostenibile? "
            "L'ottimizzatore scala proporzionalmente le spese Senior/Old/Older lasciando invariata "
            "la strategia finanziaria."
        )
        life_ages=[]
        cols=st.columns(5)
        defaults=[52,55,58,60,62]
        for i,c in enumerate(cols):
            with c:
                life_ages.append(st.number_input(
                    f"Età {i+1}",int(params["age_start"])+1,int(params["inps_age"]),
                    min(max(defaults[i],int(params["age_start"])+1),int(params["inps_age"])),
                    key=f"life_age_{i}"
                ))
        if st.button("💶 Ottimizza stile di vita", use_container_width=True):
            out_rows=[]
            prog=st.progress(0)
            for j,age in enumerate(life_ages):
                lo,hi=0.50,2.50
                best=0.0; bestm=None
                for step in range(8):
                    mid=(lo+hi)/2
                    pp=scale_expenses(params,mid); pp["retire_age"]=int(age)
                    m=mc_evaluate(pp,opt_scenario,mc_runs,seed=7200+int(age)*17+step)
                    if m["success_rate"]>=success_threshold:
                        best=mid; bestm=m; lo=mid
                    else:
                        hi=mid
                base_month=sum(params["expense_bands"]["Senior"].values())
                out_rows.append({
                    "Età uscita":int(age),
                    "Fattibile":best>0,
                    "Moltiplicatore spese":best if best>0 else np.nan,
                    "Spesa Senior max €/mese oggi":base_month*best if best>0 else np.nan,
                    "Successo MC":bestm["success_rate"] if bestm else 0.0,
                    "P10 finale":bestm["p10"] if bestm else np.nan,
                })
                prog.progress((j+1)/len(life_ages))
            ldf=pd.DataFrame(out_rows)
            st.dataframe(ldf.style.format({"Moltiplicatore spese":"{:.2f}x",
                "Spesa Senior max €/mese oggi":"€{:,.0f}","Successo MC":"{:.1%}",
                "P10 finale":"€{:,.0f}"}),use_container_width=True,hide_index=True)

    with st.expander("3 · Ottimizza Portafoglio"):
        st.write(
            "**Obiettivo:** NON massimizzare il patrimonio. Per ciascuna età scelta cerca la strategia "
            "con la **minore esposizione azionaria** che raggiunge il patrimonio finale vincolo con "
            "la probabilità Monte Carlo richiesta. Se nessuna strategia supera la soglia, restituisce NON FATTIBILE."
        )
        port_ages=[]
        cols=st.columns(5)
        defaults=[50,52,54,56,58]
        for i,c in enumerate(cols):
            with c:
                port_ages.append(st.number_input(
                    f"Pensione {i+1}",int(params["age_start"])+1,int(params["inps_age"]),
                    min(max(defaults[i],int(params["age_start"])+1),int(params["inps_age"])),
                    key=f"port_age_{i}"
                ))
        if st.button("🛡️ Ottimizza portafoglio minimo-rischio", type="primary", use_container_width=True):
            port_rows=[]
            chosen={}
            prog=st.progress(0)
            for j,age in enumerate(port_ages):
                winner=None; winner_m=None
                candidates=portfolio_candidates(params,int(age))
                for ci,cand in enumerate(candidates):
                    m=mc_evaluate(cand,opt_scenario,mc_runs,seed=99000+int(age)*101+ci)
                    if m["success_rate"]>=success_threshold:
                        winner=cand; winner_m=m
                        break
                if winner is None:
                    port_rows.append({"Età uscita":int(age),"Fattibile":False,
                        "Azionario medio vita":np.nan,"Allocazione iniziale":"—",
                        "Successo MC":0.0,"P10 finale":np.nan,"P50 finale":np.nan,
                        "Vendita azioni in cigno nero":np.nan})
                else:
                    chosen[int(age)]=winner
                    a=winner["alloc_initial"]
                    port_rows.append({"Età uscita":int(age),"Fattibile":True,
                        "Azionario medio vita":portfolio_risk_score(winner),
                        "Allocazione iniziale":f"{a[0]*100:.0f}% az. / {a[1]*100:.0f}% obbl. / {a[2]*100:.0f}% liq.",
                        "Successo MC":winner_m["success_rate"],"P10 finale":winner_m["p10"],
                        "P50 finale":winner_m["p50"],
                        "Vendita azioni in cigno nero":winner_m["crisis_sale_rate"]})
                prog.progress((j+1)/len(port_ages))
            pdf=pd.DataFrame(port_rows)
            st.dataframe(pdf.style.format({"Azionario medio vita":"{:.1%}","Successo MC":"{:.1%}",
                "P10 finale":"€{:,.0f}","P50 finale":"€{:,.0f}",
                "Vendita azioni in cigno nero":"{:.1%}"}),use_container_width=True,hide_index=True)

            if chosen:
                age_show=st.selectbox("Mostra glide path ottimizzato per età",sorted(chosen.keys()),key="show_port_age")
                win=chosen[age_show]
                gl=[{"Stadio":"Senior iniziale","Età":int(win["age_start"]),
                     "Azionario":win["alloc_initial"][0],"Obbligazioni":win["alloc_initial"][1],"Liquidità":win["alloc_initial"][2]}]
                for stg in win["allocation_stages"]:
                    gl.append({"Stadio":stg["name"],"Età":stg["age"],"Azionario":stg["alloc"][0],
                               "Obbligazioni":stg["alloc"][1],"Liquidità":stg["alloc"][2]})
                gdf=pd.DataFrame(gl)
                st.dataframe(gdf.style.format({"Azionario":"{:.0%}","Obbligazioni":"{:.0%}","Liquidità":"{:.0%}"}),
                             use_container_width=True,hide_index=True)

    st.download_button(
        "Esporta configurazione JSON",
        json.dumps(params, ensure_ascii=False, indent=2).encode("utf-8"),
        file_name="lifeplan360_config.json",
        mime="application/json",
    )

    st.download_button(
        "Scarica CSV scenario selezionato",
        results[detail_scenario].to_csv(index=False).encode("utf-8"),
        file_name=f"{detail_scenario.replace(' ', '_')}.csv",
        mime="text/csv",
    )

    excel_bytes = build_excel(params, results, summary)
    st.download_button(
        "Scarica report Excel completo",
        excel_bytes,
        file_name="LifePlan360_Report.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


st.caption(
    "LifePlan 360 è uno strumento di simulazione e pianificazione. "
    "Non sostituisce consulenza fiscale, previdenziale o finanziaria professionale."
)
